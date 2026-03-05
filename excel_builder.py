"""
Excel Builder (Korean / DART) -- creates a professional financial model workbook.

Sheet 1 "Financial Statements" : historical Income Statement, Balance Sheet,
                                   Cash Flow Statement + balance check
Sheet 2 "DCF Model"            : 5-year discounted cash flow valuation

All amounts displayed in KRW (원).
Data sourced from DART (Korean Financial Supervisory Service).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
import re
from typing import Dict, Optional, List
import statistics

# ── Palette ───────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F4E79"
MED_BLUE    = "2E75B6"
LIGHT_BLUE  = "BDD7EE"
XLIGHT_BLUE = "DEEAF1"
DARK_GREEN  = "375623"
LIGHT_GREEN = "E2EFDA"
YELLOW      = "FFF2CC"
DARK_YELLOW = "F4B942"
LIGHT_RED   = "FFE2CC"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"
MED_GRAY    = "D6D6D6"

# ── Borders ───────────────────────────────────────────────────────────────────
_thin   = Side(style='thin',   color='000000')
_medium = Side(style='medium', color='000000')
_thick  = Side(style='thick',  color='000000')
THIN_BOX  = Border(left=_thin,   right=_thin,   top=_thin,   bottom=_thin)
BOT_MED   = Border(bottom=_medium)
BOT_THICK = Border(bottom=_thick)
TOP_THIN  = Border(top=_thin)

# ── Number formats (KRW adapted) ─────────────────────────────────────────────
FMT_KRW     = '#,##0'            # e.g. 3,008,709  (KRW)
FMT_PCT     = '0.0%'
FMT_MULT    = '0.0x'
FMT_INT     = '#,##0'
FMT_EPS     = '#,##0'            # Korean EPS is usually integer (원)
FMT_SHARES  = '#,##0'            # Share count (raw)

# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def _font(bold=False, color='000000', size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic, name='Calibri')


def _align(h='left', v='center', wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _style(cell, fill_hex=None, bold=False, font_color='000000',
           h_align='left', number_format=None, border=None, italic=False):
    if fill_hex:
        cell.fill = _fill(fill_hex)
    cell.font  = _font(bold=bold, color=font_color, italic=italic)
    cell.alignment = _align(h=h_align)
    if number_format:
        cell.number_format = number_format
    if border:
        cell.border = border


def _set_col_widths(ws, widths: Dict[int, float]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _val(d: Dict[int, Optional[float]], year: int,
         scale: float = 1.0, negate: bool = False) -> Optional[float]:
    """Return value; DART values are used as-is (scale=1.0 by default).
    Set negate=True to flip sign (e.g. for expenses stored as positive)."""
    v = d.get(year)
    if v is None:
        return None
    v = v / scale
    return -v if negate else v


def _safe_avg(values: List[Optional[float]]) -> float:
    cleaned = [v for v in values if v is not None]
    return statistics.mean(cleaned) if cleaned else 0.0


def _write_section_header(ws, row: int, title: str, cols: int = 6):
    cell = ws.cell(row=row, column=1, value=title)
    _style(cell, fill_hex=MED_BLUE, bold=True, font_color=WHITE, h_align='left')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell.border = THIN_BOX


def _write_col_headers(ws, row: int, year_cols: List[int], years: List[int],
                       start_col: int = 2):
    ws.cell(row=row, column=1).value = ''
    for i, yr in enumerate(years):
        col = start_col + i
        c = ws.cell(row=row, column=col, value=f'FY{yr}')
        _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX


def _write_row(ws, row: int, label: str, data: Dict[int, Optional[float]],
               years: List[int], start_col: int = 2, scale: float = 1.0,
               fmt: str = FMT_KRW, bold: bool = False, fill: str = None,
               negate: bool = False, indent: int = 0) -> int:
    """Write one labelled data row. Returns the row number."""
    prefix = '  ' * indent
    c = ws.cell(row=row, column=1, value=prefix + label)
    _style(c, fill_hex=fill, bold=bold)

    for i, yr in enumerate(sorted(years)):
        col = start_col + i
        v = _val(data, yr, scale=scale, negate=negate)
        cell = ws.cell(row=row, column=col, value=v)
        _style(cell, fill_hex=fill or (LIGHT_GRAY if i % 2 == 0 else WHITE),
               bold=bold, h_align='right', number_format=fmt)
        if bold:
            cell.border = BOT_MED
    return row


def _spacer(ws, row: int, cols: int = 6):
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).value = None


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 -- FINANCIAL STATEMENTS (Income Statement section)
# ─────────────────────────────────────────────────────────────────────────────

def _write_financial_statements(ws, company_info, financial_data):
    """Write the Financial Statements sheet (IS / BS / CF).
    Returns fs_rows dict mapping item names to row numbers.
    """
    years_desc = financial_data['years']                # newest first
    years = list(reversed(years_desc))                  # oldest first for display
    inc  = financial_data['income_statement']
    bs   = financial_data['balance_sheet']
    cf   = financial_data['cash_flow']

    # ── Annualized column setup ──────────────────────────────────────────
    has_ann = 'annualized' in financial_data and 'ltm_info' in financial_data
    _ltm_info = financial_data.get('ltm_info', {})
    ANN_YEAR = None
    if has_ann:
        ANN_YEAR = years[-1] + 0.5          # sentinel key for annualized col
        _ann_src = financial_data['annualized']
        _ann_is = _ann_src.get('income_statement', {})
        _ann_bs = _ann_src.get('balance_sheet', {})
        _ann_cf = _ann_src.get('cash_flow', {})
        for key in inc:
            if isinstance(inc[key], dict):
                inc[key][ANN_YEAR] = _ann_is.get(key)
        for key in bs:
            if isinstance(bs[key], dict):
                bs[key][ANN_YEAR] = _ann_bs.get(key)
        for key in cf:
            if isinstance(cf[key], dict):
                cf[key][ANN_YEAR] = _ann_cf.get(key)
        years.append(ANN_YEAR)

    n    = len(years)

    # ── Projection setup ────────────────────────────────────────────────
    latest_yr      = max(yr for yr in years if isinstance(yr, int))
    proj_years     = [latest_yr + i for i in range(1, 6)]
    n_proj         = 5
    total_cols     = 1 + n + n_proj          # label + hist (+ann) + proj
    proj_start_col = 2 + n                   # first projection column

    def _cl(i: int) -> str:
        """Excel column letter for year index i (0 = oldest year = column B)."""
        return get_column_letter(2 + i)

    def _pcl(j: int) -> str:
        """Excel column letter for projection year index j (0 = first proj year)."""
        return get_column_letter(proj_start_col + j)

    def _fw(row_num: int, col_idx: int, formula: str,
            fmt: str = FMT_KRW, bold: bool = False, fill: str = None):
        """Write an Excel formula string to the data cell at (row_num, year col_idx)."""
        c = ws.cell(row=row_num, column=2 + col_idx, value=formula)
        _style(c,
               fill_hex=fill or (LIGHT_GRAY if col_idx % 2 == 0 else WHITE),
               bold=bold, h_align='right', number_format=fmt)
        if bold:
            c.border = BOT_MED
        return c

    def _pfw(row_num: int, proj_idx: int, formula: str,
             fmt: str = FMT_KRW, bold: bool = False, fill: str = None):
        """Write an Excel formula into a projection column cell."""
        c = ws.cell(row=row_num, column=proj_start_col + proj_idx, value=formula)
        _style(c,
               fill_hex=fill or (XLIGHT_BLUE if proj_idx % 2 == 0 else WHITE),
               bold=bold, h_align='right', number_format=fmt)
        if bold:
            c.border = BOT_MED
        return c

    def _lbl(row_num: int, text: str, bold: bool = False,
             fill: str = None, ind: int = 0):
        c = ws.cell(row=row_num, column=1, value='  ' * ind + text)
        _style(c, fill_hex=fill, bold=bold)
        return c

    def _fix_ann_headers(hdr_row):
        """Overwrite ltm and ann column headers with proper labels."""
        if not has_ann:
            return
        ltm_yr = _ltm_info['ltm_year']
        ltm_idx = years.index(ltm_yr)
        c = ws.cell(row=hdr_row, column=2 + ltm_idx,
                    value=_ltm_info['q_label'])
        _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE,
               h_align='center')
        c.border = THIN_BOX
        ann_idx = years.index(ANN_YEAR)
        c = ws.cell(row=hdr_row, column=2 + ann_idx,
                    value=_ltm_info['ann_label'])
        _style(c, fill_hex=DARK_GREEN, bold=True, font_color=WHITE,
               h_align='center')
        c.border = THIN_BOX

    # ── Column widths ───────────────────────────────────────────────────
    col_widths = {1: 40}
    for ci in range(2, 2 + n + n_proj):
        col_widths[ci] = 16
    _set_col_widths(ws, col_widths)

    # ── Freeze panes (freeze below row 2 so title/subtitle stay visible) ──
    ws.freeze_panes = 'A3'

    last_hist = _cl(n - 1)     # column letter of the most recent historical year

    r = 1

    # ── Title row ───────────────────────────────────────────────────────
    corp_name  = company_info.get('corp_name', '')
    stock_code = company_info.get('stock_code', '')
    title_cell = ws.cell(
        row=r, column=1,
        value=f"{corp_name}  ({stock_code})"
              f" - 재무제표 (Financial Statements)  |  단위: KRW")
    _style(title_cell, fill_hex=DARK_BLUE, bold=True, font_color=WHITE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
    ws.row_dimensions[r].height = 18
    r += 1

    # ── Subtitle row (with scenario dropdown in the last projection column) ──
    sub = ws.cell(
        row=r, column=1,
        value="출처: DART 전자공시  |  음영 행 = 수식 자동계산  |  금액 단위: KRW (원)")
    _style(sub, fill_hex=XLIGHT_BLUE, italic=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r,
                   end_column=total_cols - 1)

    # ── Scenario dropdown cell (subtitle row, last projection column) ────
    dropdown_col = proj_start_col + n_proj - 1
    dropdown_ref = f'${get_column_letter(dropdown_col)}${r}'
    dd_cell = ws.cell(row=r, column=dropdown_col, value='Base Case')
    _style(dd_cell, fill_hex=YELLOW, bold=True, h_align='center')
    dd_cell.border = THIN_BOX
    dv = DataValidation(
        type='list',
        formula1='"Best Case,Base Case,Weak Case"',
        allow_blank=False,
    )
    dv.error = 'Please select Best Case, Base Case, or Weak Case'
    dv.errorTitle = 'Invalid Scenario'
    ws.add_data_validation(dv)
    dv.add(dd_cell)
    r += 2

    # =========================================================================
    # INCOME STATEMENT  (손익계산서)
    # =========================================================================
    _write_section_header(ws, r, '손익계산서 (INCOME STATEMENT)', cols=total_cols)
    r += 1

    # Column headers -- historical (DARK_BLUE) + projection (MED_BLUE)
    _write_col_headers(ws, r, list(range(2, 2 + n)), years, start_col=2)
    _fix_ann_headers(r)
    for j, py in enumerate(proj_years):
        col = proj_start_col + j
        c = ws.cell(row=r, column=col, value=f'FY{py}E')
        _style(c, fill_hex=MED_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    r += 1

    # -- Revenue (매출액) --
    rev_row = r
    _write_row(ws, r, '매출액 (Revenue)', inc['revenue'], years)
    r += 1

    # -- COGS (매출원가) -- displayed as negative
    cogs_row = r
    _write_row(ws, r, '  매출원가 (COGS)', inc['cogs'], years, negate=True, indent=1)
    r += 1

    # Gross Profit = Revenue + COGS (COGS shown negative, so addition)  [FORMULA]
    gp_row = r
    _lbl(r, '매출총이익 (Gross Profit)', bold=True)
    for i in range(n):
        _fw(r, i, f'={_cl(i)}{rev_row}+{_cl(i)}{cogs_row}', bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{rev_row}+{_pcl(j)}{cogs_row}', bold=True, fill=XLIGHT_BLUE)
    r += 1

    # Gross Margin % = Gross Profit / Revenue  [FORMULA]
    gp_margin_row = r
    _lbl(r, '  매출총이익률 (Gross Margin %)', ind=1)
    for i in range(n):
        _fw(r, i, f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{gp_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    for j in range(n_proj):
        _pfw(r, j, f'=IF({_pcl(j)}{rev_row}<>0,{_pcl(j)}{gp_row}/{_pcl(j)}{rev_row},"")',
             fmt=FMT_PCT)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # -- R&D Expense (연구개발비) -- displayed as negative if present
    rd_row = r
    _write_row(ws, r, '  연구개발비 (R&D Expense)', inc['rd_expense'], years,
               negate=True, indent=1)
    r += 1

    # -- SGA Expense (판매비와관리비) -- displayed as negative
    sga_row = r
    _write_row(ws, r, '  판매비와관리비 (SGA Expense)', inc['sga_expense'], years,
               negate=True, indent=1)
    r += 1

    # Operating Income = Gross Profit + R&D (neg) + SGA (neg)  [FORMULA]
    oi_row = r
    _lbl(r, '영업이익 (Operating Income)', bold=True)
    for i in range(n):
        _fw(r, i,
            f'={_cl(i)}{gp_row}+{_cl(i)}{rd_row}+{_cl(i)}{sga_row}',
            bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        _pfw(r, j,
             f'={_pcl(j)}{gp_row}+{_pcl(j)}{rd_row}+{_pcl(j)}{sga_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    # Operating Margin % = Operating Income / Revenue  [FORMULA]
    oi_margin_row = r
    _lbl(r, '  영업이익률 (Operating Margin %)', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{oi_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    for j in range(n_proj):
        _pfw(r, j,
             f'=IF({_pcl(j)}{rev_row}<>0,{_pcl(j)}{oi_row}/{_pcl(j)}{rev_row},"")',
             fmt=FMT_PCT)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # -- D&A (감가상각비) --
    da_row = r
    _write_row(ws, r, '  감가상각비 (D&A)', inc['da'], years, indent=1)
    for j in range(n_proj):
        _pfw(r, j, f'={last_hist}{da_row}')
    r += 1

    # EBITDA = Operating Income + D&A  [FORMULA]
    ebitda_row = r
    _lbl(r, 'EBITDA', bold=True)
    for i in range(n):
        _fw(r, i,
            f'={_cl(i)}{oi_row}+{_cl(i)}{da_row}',
            bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        _pfw(r, j,
             f'={_pcl(j)}{oi_row}+{_pcl(j)}{da_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    # EBITDA Margin % = EBITDA / Revenue  [FORMULA]
    ebitda_margin_row = r
    _lbl(r, '  EBITDA 마진율 (EBITDA Margin %)', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{ebitda_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    for j in range(n_proj):
        _pfw(r, j,
             f'=IF({_pcl(j)}{rev_row}<>0,{_pcl(j)}{ebitda_row}/{_pcl(j)}{rev_row},"")',
             fmt=FMT_PCT)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # ── 기타손익 (Other Income/Expense) section ─────────────────────────
    _lbl(r, '기타손익 (Other Income/Expense)', bold=True, fill=LIGHT_BLUE)
    for ci in range(2, 2 + n + n_proj):
        ws.cell(row=r, column=ci).fill = _fill(LIGHT_BLUE)
    r += 1

    # Interest Expense (금융비용) -- held flat for projections
    int_exp_row = r
    _write_row(ws, r, '  금융비용 (Interest Expense)', inc['interest_expense'], years, indent=1)
    for j in range(n_proj):
        _pfw(r, j, f'={last_hist}{int_exp_row}')
    r += 1

    # Interest Income (금융수익) -- held flat for projections
    int_inc_row = r
    _write_row(ws, r, '  금융수익 (Interest Income)', inc['interest_income'], years, indent=1)
    for j in range(n_proj):
        _pfw(r, j, f'={last_hist}{int_inc_row}')
    r += 1

    # Other Income (기타수익) -- held flat for projections
    other_inc_row = r
    _write_row(ws, r, '  기타수익 (Other Income)', inc['other_income'], years, indent=1)
    for j in range(n_proj):
        _pfw(r, j, f'={last_hist}{other_inc_row}')
    r += 1

    # Pretax Income = Operating Income - Interest Expense + Interest Income + Other Income  [FORMULA]
    pretax_row = r
    _lbl(r, '법인세비용차감전순이익 (Pretax Income)', bold=True)
    for i in range(n):
        _fw(r, i,
            f'={_cl(i)}{oi_row}-{_cl(i)}{int_exp_row}+{_cl(i)}{int_inc_row}+{_cl(i)}{other_inc_row}',
            bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        _pfw(r, j,
             f'={_pcl(j)}{oi_row}-{_pcl(j)}{int_exp_row}+{_pcl(j)}{int_inc_row}+{_pcl(j)}{other_inc_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    # Tax Expense (법인세비용) -- projection backfilled after assumptions band
    tax_row = r
    _write_row(ws, r, '  법인세비용 (Tax Expense)', inc['tax_expense'], years, indent=1)
    r += 1

    # Effective Tax Rate % = Tax / Pretax  [FORMULA]
    eff_tax_row = r
    _lbl(r, '  유효법인세율 (Effective Tax Rate %)', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{pretax_row}<>0,{_cl(i)}{tax_row}/{_cl(i)}{pretax_row},"")',
            fmt=FMT_PCT)
    for j in range(n_proj):
        _pfw(r, j,
             f'=IF({_pcl(j)}{pretax_row}<>0,{_pcl(j)}{tax_row}/{_pcl(j)}{pretax_row},"")',
             fmt=FMT_PCT)
    r += 1

    # Net Income = Pretax Income - Tax Expense  [FORMULA]
    ni_row = r
    _lbl(r, '당기순이익 (Net Income)', bold=True)
    for i in range(n):
        _fw(r, i, f'={_cl(i)}{pretax_row}-{_cl(i)}{tax_row}', bold=True, fill=LIGHT_GREEN)
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{pretax_row}-{_pcl(j)}{tax_row}', bold=True, fill=LIGHT_GREEN)
    r += 1

    # Net Margin % = Net Income / Revenue  [FORMULA]
    net_margin_row = r
    _lbl(r, '  순이익률 (Net Margin %)', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{ni_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    for j in range(n_proj):
        _pfw(r, j,
             f'=IF({_pcl(j)}{rev_row}<>0,{_pcl(j)}{ni_row}/{_pcl(j)}{rev_row},"")',
             fmt=FMT_PCT)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # ── EPS / Shares section ────────────────────────────────────────────
    _write_row(ws, r, '  기본주당이익 (EPS Basic)', inc['eps_basic'], years,
               scale=1.0, fmt=FMT_EPS, indent=1)
    r += 1

    eps_diluted_row = r
    _write_row(ws, r, '  희석주당이익 (EPS Diluted)', inc['eps_diluted'], years,
               scale=1.0, fmt=FMT_EPS, indent=1)
    r += 1

    _write_row(ws, r, '  보통주식수 (Shares Basic)', inc['shares_basic'], years,
               scale=1.0, fmt=FMT_SHARES, indent=1)
    r += 1

    shares_diluted_row = r
    _write_row(ws, r, '  희석주식수 (Shares Diluted)', inc['shares_diluted'], years,
               scale=1.0, fmt=FMT_SHARES, indent=1)
    r += 1

    r += 1  # blank row before assumptions

    # =========================================================================
    # 수익성 가정 (Profitability Assumptions)
    # =========================================================================
    _lbl(r, '수익성 가정 (Profitability Assumptions)', bold=True, fill=LIGHT_BLUE)
    for ci in range(2, 2 + n + n_proj):
        ws.cell(row=r, column=ci).fill = _fill(LIGHT_BLUE)
    r += 1

    # Revenue Growth (YoY %) -- N/A for oldest year
    rev_growth_asm_row = r
    _lbl(r, '  매출액 성장률 (Revenue Growth %)', ind=1)
    for i in range(n):
        if i == 0:
            c = ws.cell(row=r, column=2 + i, value='N/A')
            _style(c, fill_hex=MED_GRAY, h_align='center', italic=True)
        elif has_ann and years[i] == _ltm_info.get('ltm_year'):
            # Q3 cumulative: 9mo vs 12mo not comparable
            c = ws.cell(row=r, column=2 + i, value='N/A')
            _style(c, fill_hex=MED_GRAY, h_align='center', italic=True)
        elif has_ann and years[i] == ANN_YEAR:
            # Annualized growth vs latest actual annual year
            base_idx = years.index(_ltm_info['base_year'])
            _fw(r, i,
                f'=IF({_cl(base_idx)}{rev_row}<>0,{_cl(i)}{rev_row}/{_cl(base_idx)}{rev_row}-1,"")',
                fmt=FMT_PCT)
        else:
            _fw(r, i,
                f'=IF({_cl(i-1)}{rev_row}<>0,{_cl(i)}{rev_row}/{_cl(i-1)}{rev_row}-1,"")',
                fmt=FMT_PCT)
    r += 1

    # Gross Margin (%)
    gp_margin_asm_row = r
    _lbl(r, '  매출총이익률 (Gross Margin %)', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{gp_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    r += 1

    # SGA % of Revenue
    sga_pct_asm_row = r
    _lbl(r, '  판관비 비율 (SGA % of Revenue)', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,ABS({_cl(i)}{sga_row})/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    r += 1

    # Operating Margin (%)
    oi_margin_asm_row = r
    _lbl(r, '  영업이익률 (Operating Margin %)', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{oi_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    r += 1

    # Tax Rate (%)
    tax_rate_asm_row = r
    _lbl(r, '  법인세율 (Tax Rate %)', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{pretax_row}<>0,{_cl(i)}{tax_row}/{_cl(i)}{pretax_row},"")',
            fmt=FMT_PCT)
    r += 1

    # D&A % of Revenue
    da_pct_asm_row = r
    _lbl(r, '  감가상각비 비율 (D&A % of Revenue)', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{da_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # ── Backfill IS projection formulas ─────────────────────────────
    # Now that assumptions band row numbers are known, fill in the
    # projection columns for the IS line items above.
    # NOTE: Assumption projection columns will be filled later by
    # the scenario table section (IF(dropdown) formulas referencing
    # Best/Base/Weak case rows). For now we use a simple approach:
    # hold historical averages or last-year values as placeholders.
    # The full scenario table will be implemented in a later section.

    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        # Revenue = prior year * (1 + growth assumption)
        _pfw(rev_row, j, f'={prev}{rev_row}*(1+{_pcl(j)}{rev_growth_asm_row})')
        # COGS = -(Revenue * (1 - GP margin assumption))
        _pfw(cogs_row, j, f'=-{_pcl(j)}{rev_row}*(1-{_pcl(j)}{gp_margin_asm_row})')
        # R&D: hold flat from last historical
        _pfw(rd_row, j, f'={last_hist}{rd_row}')
        # SGA = -(Revenue * SGA% assumption)
        _pfw(sga_row, j, f'=-{_pcl(j)}{rev_row}*{_pcl(j)}{sga_pct_asm_row}')
        # Tax = Pretax * Tax Rate assumption
        _pfw(tax_row, j, f'={_pcl(j)}{pretax_row}*{_pcl(j)}{tax_rate_asm_row}')

    # ── Compute base-case values from historical averages for scenario table ──
    def _hist_vals(num_data, denom_data, as_growth=False):
        """Compute historical ratios or growth rates from raw data.
        Excludes Q3 cumulative and annualized sentinel years."""
        vals = []
        # Use only actual full-year annual data for averages
        _excl = set()
        if has_ann:
            _excl.add(_ltm_info.get('ltm_year'))
            _excl.add(ANN_YEAR)
        sorted_yrs = sorted(yr for yr in years if yr not in _excl)
        if as_growth:
            for idx in range(1, len(sorted_yrs)):
                y0, y1 = sorted_yrs[idx - 1], sorted_yrs[idx]
                v0 = num_data.get(y0)
                v1 = num_data.get(y1)
                if v0 and v1 and v0 > 0:
                    vals.append(v1 / v0 - 1.0)
        else:
            for yr in sorted_yrs:
                nv = num_data.get(yr)
                dv = denom_data.get(yr) if denom_data else None
                if nv is not None and dv and dv != 0:
                    vals.append(nv / dv)
        return _safe_avg(vals) if vals else None

    avg_rev_growth = _hist_vals(inc['revenue'], None, as_growth=True) or 0.05
    avg_gp_margin  = _hist_vals(
        {yr: (inc['revenue'].get(yr) or 0) - (inc['cogs'].get(yr) or 0) for yr in years},
        inc['revenue']) or 0.35
    avg_sga_pct    = _hist_vals(inc['sga_expense'], inc['revenue']) or 0.10
    avg_oi_margin  = _hist_vals(inc['operating_income'], inc['revenue']) or 0.10
    avg_tax_rate   = _hist_vals(inc['tax_expense'], inc['pretax_income']) or 0.22
    avg_da_pct     = _hist_vals(inc['da'], inc['revenue']) or 0.05

    # =========================================================================
    # SCENARIO ASSUMPTIONS TABLE  (시나리오 가정표)
    # Three cases (Best / Base / Weak) for each projection driver.
    # The dropdown cell in the title area selects which case to use.
    # =========================================================================
    r += 2
    _write_section_header(
        ws, r,
        '시나리오 가정 (Scenario Assumptions)  |  노란색 셀을 수정하여 추정치를 조정하세요',
        cols=total_cols)
    r += 1

    # Column headers for projection years + delta column
    delta_col = proj_start_col - 1               # last historical column
    delta_cl  = get_column_letter(delta_col)
    ws.cell(row=r, column=1, value='').fill = _fill(DARK_BLUE)
    for ci in range(2, delta_col):
        ws.cell(row=r, column=ci).fill = _fill(DARK_BLUE)
    c = ws.cell(row=r, column=delta_col, value='+/\u2212')
    _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE, h_align='center')
    c.border = THIN_BOX
    for j, py in enumerate(proj_years):
        col = proj_start_col + j
        c = ws.cell(row=r, column=col, value=f'FY{py}E')
        _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    r += 1

    # Define metrics: (key, label, base_val, spread, invert)
    # invert=True means lower is "better" (costs / tax)
    metrics_def = [
        ('rev_growth',     '매출 성장률 (Revenue Growth %)',     avg_rev_growth,     0.03, False),
        ('gp_margin',      '매출총이익률 (Gross Margin %)',      avg_gp_margin,      0.03, False),
        ('sga_pct',        '판관비 비율 (SGA % of Revenue)',     avg_sga_pct,        0.02, True),
        ('oi_margin',      '영업이익률 (Operating Margin %)',    avg_oi_margin,      0.03, False),
        ('tax_rate',       '법인세율 (Tax Rate %)',              avg_tax_rate,       0.03, True),
        ('da_pct',         '감가상각비 비율 (D&A % of Revenue)', avg_da_pct,         0.01, True),
    ]

    FMT_DELTA = '+0.0%;-0.0%;0.0%'

    scenario_rows = {}  # {metric_key: {'best': row, 'base': row, 'weak': row}}

    for metric_key, metric_label, base_val, spread, invert in metrics_def:
        # Sub-header for this metric
        _lbl(r, metric_label, bold=True, fill=LIGHT_BLUE)
        for ci in range(2, 2 + n + n_proj):
            ws.cell(row=r, column=ci).fill = _fill(LIGHT_BLUE)
        r += 1

        # Signed deltas: for "invert" metrics (costs), best = negative spread
        best_delta = -spread if invert else +spread
        weak_delta = +spread if invert else -spread

        # Row assignments (Best / Base / Weak in consecutive rows)
        best_row = r
        base_row = r + 1
        weak_row = r + 2

        scenario_rows[metric_key] = {
            'best': best_row,
            'base': base_row,
            'weak': weak_row,
        }

        # ── Best Case: delta cell + formula cells ──
        _lbl(best_row, '  Best Case (최적)', ind=1)
        c = ws.cell(row=best_row, column=delta_col, value=round(best_delta, 4))
        _style(c, fill_hex=YELLOW, bold=False, h_align='right',
               number_format=FMT_DELTA)
        c.border = THIN_BOX
        for j in range(n_proj):
            col = proj_start_col + j
            cl_letter = get_column_letter(col)
            c = ws.cell(row=best_row, column=col,
                        value=f'={cl_letter}{base_row}+{delta_cl}{best_row}')
            _style(c, fill_hex=LIGHT_GREEN, bold=False, h_align='right',
                   number_format=FMT_PCT)
            c.border = THIN_BOX

        # ── Base Case ──
        _lbl(base_row, '  Base Case (기본)', ind=1)
        for j in range(n_proj):
            col = proj_start_col + j
            c = ws.cell(row=base_row, column=col, value=round(base_val, 4))
            _style(c, fill_hex=YELLOW, bold=False, h_align='right',
                   number_format=FMT_PCT)
            c.border = THIN_BOX

        # ── Weak Case: delta cell + formula cells ──
        _lbl(weak_row, '  Weak Case (약세)', ind=1)
        c = ws.cell(row=weak_row, column=delta_col, value=round(weak_delta, 4))
        _style(c, fill_hex=YELLOW, bold=False, h_align='right',
               number_format=FMT_DELTA)
        c.border = THIN_BOX
        for j in range(n_proj):
            col = proj_start_col + j
            cl_letter = get_column_letter(col)
            c = ws.cell(row=weak_row, column=col,
                        value=f'={cl_letter}{base_row}+{delta_cl}{weak_row}')
            _style(c, fill_hex=LIGHT_GREEN, bold=False, h_align='right',
                   number_format=FMT_PCT)
            c.border = THIN_BOX

        r = weak_row + 1
        r += 1   # spacer between metrics

    # ── Deferred fill: assumptions band projection IF formulas ────────
    # Now that scenario table row numbers are known, fill projection
    # columns of the assumptions band with IF(dropdown) formulas.
    asm_band_metrics = [
        (rev_growth_asm_row,    'rev_growth'),
        (gp_margin_asm_row,     'gp_margin'),
        (sga_pct_asm_row,       'sga_pct'),
        (oi_margin_asm_row,     'oi_margin'),
        (tax_rate_asm_row,      'tax_rate'),
        (da_pct_asm_row,        'da_pct'),
    ]
    for j in range(n_proj):
        col = proj_start_col + j
        col_letter = get_column_letter(col)
        for asm_row, metric_key in asm_band_metrics:
            best_ref = f'{col_letter}{scenario_rows[metric_key]["best"]}'
            base_ref = f'{col_letter}{scenario_rows[metric_key]["base"]}'
            weak_ref = f'{col_letter}{scenario_rows[metric_key]["weak"]}'
            formula = (
                f'=IF({dropdown_ref}="Best Case",{best_ref},'
                f'IF({dropdown_ref}="Weak Case",{weak_ref},{base_ref}))'
            )
            c = ws.cell(row=asm_row, column=col, value=formula)
            _style(c, fill_hex=LIGHT_GREEN, bold=False, h_align='right',
                   number_format=FMT_PCT)
            c.border = THIN_BOX

    # =========================================================================
    # BALANCE SHEET  (재무상태표)  (historical + projections)
    # =========================================================================
    _write_section_header(ws, r, '재무상태표 (BALANCE SHEET)', cols=total_cols);  r += 1
    _write_col_headers(ws, r, list(range(2, 2 + n)), years, start_col=2)
    _fix_ann_headers(r)
    for j, py in enumerate(proj_years):
        col = proj_start_col + j
        c = ws.cell(row=r, column=col, value=f'FY{py}E')
        _style(c, fill_hex=MED_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    r += 1

    # Helper: growth-driven projection (prior year * (1 + rev growth))
    def _bs_grow(row_num):
        for j in range(n_proj):
            prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
            _pfw(row_num, j, f'={prev}{row_num}*(1+{_pcl(j)}{rev_growth_asm_row})')

    # Helper: hold-flat projection (= last historical year)
    def _bs_flat(row_num):
        for j in range(n_proj):
            _pfw(row_num, j, f'={last_hist}{row_num}')

    _write_row(ws, r, '자산 (ASSETS)', {y: None for y in years}, years, bold=True); r += 1

    # Cash -- backfilled later (will reference ending cash from CF recon)
    cash_row = r
    _write_row(ws, r, '  현금및현금성자산 (Cash)', bs['cash'], years, indent=1)
    _bs_flat(r)  # temporary: hold flat; will be overwritten after CF projections
    r += 1

    # ST Investments -- grow with revenue
    st_inv_row = r
    _write_row(ws, r, '  단기금융상품 (ST Investments)', bs['st_investments'], years, indent=1)
    _bs_grow(r); r += 1

    # Accounts Receivable -- grow with revenue
    ar_row = r
    _write_row(ws, r, '  매출채권 (Accounts Receivable)', bs['accounts_rec'], years, indent=1)
    _bs_grow(r); r += 1

    # Inventory -- grows with COGS ratio
    inventory_row = r
    _write_row(ws, r, '  재고자산 (Inventory)', bs['inventory'], years, indent=1)
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        _pfw(r, j,
             f'=IF({prev}{cogs_row}<>0,'
             f'{prev}{inventory_row}*{_pcl(j)}{cogs_row}/{prev}{cogs_row},'
             f'{prev}{inventory_row})')
    r += 1

    # Other Current Assets (plug) -- grow with revenue
    other_ca_plug = {}
    for yr in years:
        tca = bs['total_current_a'].get(yr)
        if tca is not None:
            other_ca_plug[yr] = tca - sum(
                bs[k].get(yr) or 0
                for k in ('cash', 'st_investments', 'accounts_rec', 'inventory'))
        else:
            other_ca_plug[yr] = None
    other_ca_row = r
    _write_row(ws, r, '  기타유동자산 (Other Current Assets)', other_ca_plug, years, indent=1)
    ws.cell(row=r, column=1).comment = Comment(
        "잔여: 유동자산합계 - 현금 - 단기금융상품"
        " - 매출채권 - 재고자산.\n\n"
        "포함 가능 항목:\n"
        "- 선급비용\n"
        "- 이연법인세자산 (유동)\n"
        "- 기타수취채권\n"
        "- 매각예정자산\n"
        "- 계약자산",
        "Financial Model")
    _bs_grow(r); r += 1

    # Total Current Assets [FORMULA for projections]
    total_ca_row = r
    _write_row(ws, r, '유동자산 (Total Current Assets)', bs['total_current_a'], years, bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{cash_row}+{cl}{st_inv_row}+{cl}{ar_row}'
             f'+{cl}{inventory_row}+{cl}{other_ca_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # PP&E -- backfilled after PP&E schedule is written
    ppe_row = r
    _write_row(ws, r, '  유형자산 (PP&E Net)', bs['ppe_net'], years, indent=1)
    _bs_flat(r)  # temporary: overwritten after PP&E schedule
    r += 1

    # Goodwill -- grow with revenue
    gw_row = r
    _write_row(ws, r, '  영업권 (Goodwill)', bs['goodwill'], years, indent=1)
    _bs_grow(r); r += 1

    # Intangible Assets -- grow with revenue
    intangibles_row = r
    _write_row(ws, r, '  무형자산 (Intangibles)', bs['intangibles'], years, indent=1)
    _bs_grow(r); r += 1

    # LT Investments -- grow with revenue
    lt_inv_row = r
    _write_row(ws, r, '  장기투자자산 (LT Investments)', bs['lt_investments'], years, indent=1)
    _bs_grow(r); r += 1

    # Other Non-Current Assets (plug) -- grow with revenue
    other_nca_plug = {}
    for yr in years:
        ta = bs['total_assets'].get(yr)
        tca = bs['total_current_a'].get(yr)
        if ta is not None and tca is not None:
            other_nca_plug[yr] = ta - tca - sum(
                bs[k].get(yr) or 0
                for k in ('ppe_net', 'goodwill', 'intangibles', 'lt_investments'))
        else:
            other_nca_plug[yr] = None
    other_nca_row = r
    _write_row(ws, r, '  기타비유동자산 (Other Non-Current Assets)', other_nca_plug, years, indent=1)
    ws.cell(row=r, column=1).comment = Comment(
        "잔여: 자산총계 - 유동자산 - 유형자산"
        " - 영업권 - 무형자산 - 장기투자자산.\n\n"
        "포함 가능 항목:\n"
        "- 사용권자산\n"
        "- 이연법인세자산 (비유동)\n"
        "- 비유동계약자산\n"
        "- 기타장기자산",
        "Financial Model")
    _bs_grow(r); r += 1

    # Total Assets [FORMULA for projections]
    total_assets_row = r
    _write_row(ws, r, '자산총계 (Total Assets)', bs['total_assets'], years, bold=True, fill=LIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{total_ca_row}+{cl}{ppe_row}+{cl}{gw_row}'
             f'+{cl}{intangibles_row}+{cl}{lt_inv_row}+{cl}{other_nca_row}',
             bold=True, fill=LIGHT_BLUE)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # -- Liabilities --
    _write_row(ws, r, '부채 (LIABILITIES)', {y: None for y in years}, years, bold=True); r += 1

    ap_row = r
    _write_row(ws, r, '  매입채무 (Accounts Payable)', bs['accounts_pay'], years, indent=1)
    # Projection: prior year AP * (this year COGS / prior year COGS)
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        _pfw(r, j,
             f'=IF({prev}{cogs_row}<>0,'
             f'{prev}{ap_row}*{_pcl(j)}{cogs_row}/{prev}{cogs_row},'
             f'{prev}{ap_row})')
    r += 1

    accrued_row = r
    _write_row(ws, r, '  미지급비용 (Accrued Liabilities)', bs['accrued_liab'], years, indent=1)
    _bs_grow(r); r += 1

    st_debt_row = r
    _write_row(ws, r, '  단기차입금 (Short-term Debt)', bs['st_debt'], years, indent=1)
    _bs_grow(r); r += 1

    deferred_rev_row = r
    _write_row(ws, r, '  선수금 (Deferred Revenue)', bs['deferred_rev_cur'], years, indent=1)
    _bs_grow(r); r += 1

    # Other Current Liabilities (plug) -- grow with revenue
    other_cl_plug = {}
    for yr in years:
        tcl = bs['total_current_l'].get(yr)
        if tcl is not None:
            other_cl_plug[yr] = tcl - sum(
                bs[k].get(yr) or 0
                for k in ('accounts_pay', 'accrued_liab', 'st_debt', 'deferred_rev_cur'))
        else:
            other_cl_plug[yr] = None
    other_cl_row = r
    _write_row(ws, r, '  기타유동부채 (Other Current Liabilities)', other_cl_plug, years, indent=1)
    ws.cell(row=r, column=1).comment = Comment(
        "잔여: 유동부채합계 - 매입채무"
        " - 미지급비용 - 단기차입금 - 선수금.\n\n"
        "포함 가능 항목:\n"
        "- 리스부채 (유동)\n"
        "- 미지급법인세\n"
        "- 미지급배당금\n"
        "- 기타유동부채",
        "Financial Model")
    _bs_grow(r); r += 1

    # Total Current Liabilities [FORMULA for projections]
    total_cl_row = r
    _write_row(ws, r, '유동부채 (Total Current Liabilities)', bs['total_current_l'], years, bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{ap_row}+{cl}{accrued_row}+{cl}{st_debt_row}'
             f'+{cl}{deferred_rev_row}+{cl}{other_cl_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    lt_debt_row = r
    _write_row(ws, r, '  장기차입금 (Long-term Debt)', bs['lt_debt'], years, indent=1)
    _bs_grow(r); r += 1

    dtl_row = r
    _write_row(ws, r, '  이연법인세부채 (Deferred Tax Liab.)', bs['deferred_tax_l'], years, indent=1)
    _bs_grow(r); r += 1

    # Other Non-Current Liabilities (plug) -- grow with revenue
    other_ncl_plug = {}
    for yr in years:
        tl = bs['total_liabilities'].get(yr)
        tcl = bs['total_current_l'].get(yr)
        if tl is not None and tcl is not None:
            other_ncl_plug[yr] = tl - tcl - sum(
                bs[k].get(yr) or 0
                for k in ('lt_debt', 'deferred_tax_l'))
        else:
            other_ncl_plug[yr] = None
    other_ncl_row = r
    _write_row(ws, r, '  기타비유동부채 (Other Non-Current Liabilities)', other_ncl_plug, years, indent=1)
    ws.cell(row=r, column=1).comment = Comment(
        "잔여: 부채총계 - 유동부채합계"
        " - 장기차입금 - 이연법인세부채.\n\n"
        "포함 가능 항목:\n"
        "- 리스부채 (비유동)\n"
        "- 퇴직급여충당부채\n"
        "- 불확실한 세무포지션\n"
        "- 비유동선수금\n"
        "- 기타비유동부채",
        "Financial Model")
    _bs_grow(r); r += 1

    # Total Liabilities [FORMULA for projections]
    total_liabilities_row = r
    _write_row(ws, r, '부채총계 (Total Liabilities)', bs['total_liabilities'], years, bold=True, fill=LIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{total_cl_row}+{cl}{lt_debt_row}+{cl}{dtl_row}+{cl}{other_ncl_row}',
             bold=True, fill=LIGHT_BLUE)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # -- Shareholders' Equity --
    _write_row(ws, r, '자본 (EQUITY)', {y: None for y in years}, years, bold=True); r += 1

    cs_row = r
    _write_row(ws, r, '  자본금 (Common Stock)', bs['common_stock'], years, indent=1)
    _bs_flat(r); r += 1

    apic_row = r
    _write_row(ws, r, '  주식발행초과금 (APIC)', bs['apic'], years, indent=1)
    _bs_flat(r); r += 1

    re_row = r
    _write_row(ws, r, '  이익잉여금 (Retained Earnings)', bs['retained_earnings'], years, indent=1)
    # Projection: prior year RE + Net Income (backfilled after RE schedule)
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        _pfw(r, j, f'={prev}{re_row}+{_pcl(j)}{ni_row}')
    r += 1

    ts_row = r
    _write_row(ws, r, '  기타자본항목 (Treasury Stock / Other Capital)', bs['treasury_stock'], years, indent=1)
    _bs_flat(r); r += 1

    # Total Equity = Total Assets - Total Liabilities [FORMULA for projections]
    total_equity_row = r
    _write_row(ws, r, '자본총계 (Total Equity)', bs['total_equity'], years, bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j, f'={cl}{total_assets_row}-{cl}{total_liabilities_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # Total Liabilities + Equity [FORMULA]
    _lbl(r, '부채와자본총계 (Total Liabilities & Equity)', bold=True)
    for i in range(n):
        _fw(r, i,
            f'={_cl(i)}{total_liabilities_row}+{_cl(i)}{total_equity_row}',
            bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j, f'={cl}{total_liabilities_row}+{cl}{total_equity_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    # Balance Check: 0 = balanced, FALSE = imbalance
    _lbl(r, '대차검증 (Balance Check: 0 = OK | FALSE = 불일치)', bold=True)
    for i in range(n):
        a_cell = f'{_cl(i)}{total_assets_row}'
        diff   = f'ABS({a_cell}-({_cl(i)}{total_liabilities_row}+{_cl(i)}{total_equity_row}))'
        tol    = f'MAX({a_cell}*0.00001,0.5)'
        formula = f'=IF({diff}<{tol},0,FALSE)'
        c = ws.cell(row=r, column=2 + i, value=formula)
        _style(c, fill_hex=YELLOW, bold=True, h_align='center')
    # Projection balance check: always 0 by construction (TE = TA - TL)
    for j in range(n_proj):
        c = ws.cell(row=r, column=proj_start_col + j, value=0)
        _style(c, fill_hex=YELLOW, bold=True, h_align='center',
               number_format=FMT_KRW)
    r += 2

    # =========================================================================
    # CASH FLOW STATEMENT  (현금흐름표)  (historical + projections)
    # =========================================================================
    cf_cols = total_cols
    _write_section_header(ws, r, '현금흐름표 (CASH FLOW STATEMENT)', cols=cf_cols);  r += 1
    _write_col_headers(ws, r, list(range(2, 2 + n)), years, start_col=2)
    _fix_ann_headers(r)
    for j, py in enumerate(proj_years):
        col = proj_start_col + j
        c = ws.cell(row=r, column=col, value=f'FY{py}E')
        _style(c, fill_hex=MED_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    r += 1

    _write_row(ws, r, '영업활동 (OPERATING ACTIVITIES)', {y: None for y in years}, years, bold=True); r += 1

    # -- Net Income --
    cf_ni_row = r
    _write_row(ws, r, '  당기순이익 (Net Income)', cf['net_income'], years, indent=1)
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{ni_row}')
    r += 1

    # -- D&A --
    cf_da_row = r
    _write_row(ws, r, '  감가상각비 (D&A)', cf['da'], years, indent=1)
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{da_row}')
    r += 1

    # -- SBC (may be all None for Korean companies) --
    cf_sbc_row = r
    _write_row(ws, r, '  주식보상비용 (SBC)', cf['sbc'], years, indent=1)
    for j in range(n_proj):
        _pfw(r, j, f'=0')
    r += 1

    # -- Decreases / (Increases) in Working Capital Assets --
    cf_wc_a_row = r
    _lbl(r, '  운전자본자산 증감 (WC Assets Change)', ind=1)
    for i in range(n):
        if i == 0:
            c = ws.cell(row=r, column=2 + i, value=0)
            _style(c, fill_hex=MED_GRAY, h_align='right', number_format=FMT_KRW,
                   italic=True)
        elif has_ann and years[i] == ANN_YEAR:
            # Ann WC change: base_year → Q3 BS (same as Q3 col)
            base_idx = years.index(_ltm_info['base_year'])
            prev = _cl(base_idx); curr = _cl(i)
            _fw(r, i,
                f'=({prev}{ar_row}+{prev}{inventory_row}+{prev}{other_ca_row})'
                f'-({curr}{ar_row}+{curr}{inventory_row}+{curr}{other_ca_row})')
        else:
            prev = _cl(i - 1); curr = _cl(i)
            _fw(r, i,
                f'=({prev}{ar_row}+{prev}{inventory_row}+{prev}{other_ca_row})'
                f'-({curr}{ar_row}+{curr}{inventory_row}+{curr}{other_ca_row})')
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        curr = _pcl(j)
        _pfw(r, j,
             f'=({prev}{ar_row}+{prev}{inventory_row}+{prev}{other_ca_row})'
             f'-({curr}{ar_row}+{curr}{inventory_row}+{curr}{other_ca_row})')
    r += 1

    # -- Increases / (Decreases) in Working Capital Liabilities --
    cf_wc_l_row = r
    _lbl(r, '  운전자본부채 증감 (WC Liabilities Change)', ind=1)
    for i in range(n):
        if i == 0:
            c = ws.cell(row=r, column=2 + i, value=0)
            _style(c, fill_hex=MED_GRAY, h_align='right', number_format=FMT_KRW,
                   italic=True)
        elif has_ann and years[i] == ANN_YEAR:
            base_idx = years.index(_ltm_info['base_year'])
            prev = _cl(base_idx); curr = _cl(i)
            _fw(r, i,
                f'=({curr}{ap_row}+{curr}{accrued_row}+{curr}{deferred_rev_row}+{curr}{other_cl_row})'
                f'-({prev}{ap_row}+{prev}{accrued_row}+{prev}{deferred_rev_row}+{prev}{other_cl_row})')
        else:
            prev = _cl(i - 1); curr = _cl(i)
            _fw(r, i,
                f'=({curr}{ap_row}+{curr}{accrued_row}+{curr}{deferred_rev_row}+{curr}{other_cl_row})'
                f'-({prev}{ap_row}+{prev}{accrued_row}+{prev}{deferred_rev_row}+{prev}{other_cl_row})')
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        curr = _pcl(j)
        _pfw(r, j,
             f'=({curr}{ap_row}+{curr}{accrued_row}+{curr}{deferred_rev_row}+{curr}{other_cl_row})'
             f'-({prev}{ap_row}+{prev}{accrued_row}+{prev}{deferred_rev_row}+{prev}{other_cl_row})')
    r += 1

    # -- Other Operating Activities (plug) --
    cf_other_op_row = r
    _lbl(r, '  기타영업활동 (Other Operating, plug)', ind=1)
    op_cf_row = r + 1
    for i in range(n):
        cl_i = _cl(i)
        _fw(r, i,
            f'={cl_i}{op_cf_row}-{cl_i}{cf_ni_row}-{cl_i}{cf_da_row}'
            f'-{cl_i}{cf_sbc_row}-{cl_i}{cf_wc_a_row}-{cl_i}{cf_wc_l_row}')
    ws.cell(row=r, column=1).comment = Comment(
        "잔여: 영업활동현금흐름 - 당기순이익 - 감가상각비 - SBC"
        " - 운전자본자산 증감 - 운전자본부채 증감.\n\n"
        "포함 가능 항목:\n"
        "- 이연법인세\n"
        "- 유무형자산 손상차손\n"
        "- 투자손익\n"
        "- 기타비현금 조정항목",
        "Financial Model")
    for j in range(n_proj):
        _pfw(r, j, f'=0')
    r += 1

    # Cash from Operations
    _write_row(ws, r, '영업활동현금흐름 (Operating CF)', cf['operating_cf'], years,
               bold=True, fill=LIGHT_GREEN)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{cf_ni_row}+{cl}{cf_da_row}+{cl}{cf_sbc_row}'
             f'+{cl}{cf_wc_a_row}+{cl}{cf_wc_l_row}+{cl}{cf_other_op_row}',
             bold=True, fill=LIGHT_GREEN)
    r += 1

    _spacer(ws, r, cf_cols); r += 1

    # -- INVESTING ACTIVITIES --
    _write_row(ws, r, '투자활동 (INVESTING ACTIVITIES)', {y: None for y in years}, years, bold=True); r += 1
    capex_row = r
    _write_row(ws, r, '  유형자산취득 (Capex)', cf['capex'], years,
               negate=True, indent=1)
    for j in range(n_proj):
        _pfw(r, j, f'=0')  # placeholder -- backfilled after PP&E schedule
    r += 1

    _write_row(ws, r, '  인수합병 (Acquisitions)', cf['acquisitions'], years,
               negate=True, indent=1)
    for j in range(n_proj):
        _pfw(r, j, f'=0')
    r += 1

    # Other Investing Activities: plug for historical, 0 for projections
    other_inv_plug = {}
    for yr in years:
        inv = cf['investing_cf'].get(yr)
        if inv is not None:
            other_inv_plug[yr] = inv + sum(
                cf[k].get(yr) or 0
                for k in ('capex', 'acquisitions'))
        else:
            other_inv_plug[yr] = None
    _write_row(ws, r, '  기타투자활동 (Other Investing, plug)', other_inv_plug,
               years, indent=1)
    ws.cell(row=r, column=1).comment = Comment(
        "잔여: 투자활동현금흐름 + 유형자산취득 + 인수합병.\n\n"
        "포함 가능 항목:\n"
        "- 유가증권 취득/처분\n"
        "- 비시장성 유가증권 취득/처분\n"
        "- 기타투자활동",
        "Financial Model")
    for j in range(n_proj):
        _pfw(r, j, f'=0')
    r += 1

    inv_cf_row = r
    _write_row(ws, r, '투자활동현금흐름 (Investing CF)', cf['investing_cf'], years,
               bold=True, fill=LIGHT_RED)
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{capex_row}', bold=True, fill=LIGHT_RED)
    r += 1

    _spacer(ws, r, cf_cols); r += 1

    # -- FINANCING ACTIVITIES --
    _write_row(ws, r, '재무활동 (FINANCING ACTIVITIES)', {y: None for y in years}, years, bold=True); r += 1

    dividends_row = r
    _write_row(ws, r, '  배당금지급 (Dividends)', cf['dividends'], years, negate=True, indent=1)
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        cl = _pcl(j)
        _pfw(r, j, f'=-MIN(ABS({prev}{dividends_row}),ABS({cl}{ni_row}))')
    r += 1

    repurchases_row = r
    _write_row(ws, r, '  자사주매입 (Repurchases)', cf['repurchases'], years, negate=True, indent=1)
    for j in range(n_proj):
        _pfw(r, j, f'=0')  # placeholder, backfilled after RE schedule
    r += 1

    debt_iss_row = r
    _write_row(ws, r, '  차입금증가 (Debt Issuance)', cf['debt_issuance'], years, indent=1)
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        cl = _pcl(j)
        _pfw(r, j, f'={prev}{debt_iss_row}*(1+{cl}{rev_growth_asm_row})')
    r += 1

    debt_rep_row = r
    _write_row(ws, r, '  차입금상환 (Debt Repayment)', cf['debt_repay'], years, negate=True, indent=1)
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        cl = _pcl(j)
        _pfw(r, j,
             f'=-MIN(ABS({prev}{debt_rep_row})*(1+{cl}{rev_growth_asm_row}),'
             f'{cl}{st_debt_row}+{cl}{lt_debt_row})')
    r += 1

    # Other Financing Activities (plug)
    other_fin_plug = {}
    for yr in years:
        fin = cf['financing_cf'].get(yr)
        if fin is not None:
            other_fin_plug[yr] = (fin
                + (cf['dividends'].get(yr) or 0)
                + (cf['repurchases'].get(yr) or 0)
                - (cf['debt_issuance'].get(yr) or 0)
                + (cf['debt_repay'].get(yr) or 0))
        else:
            other_fin_plug[yr] = None
    other_fin_row = r
    _write_row(ws, r, '  기타재무활동 (Other Financing, plug)', other_fin_plug,
               years, indent=1)
    ws.cell(row=r, column=1).comment = Comment(
        "잔여: 재무활동현금흐름 + 배당금 + 자사주매입"
        " - 차입금증가 + 차입금상환.\n\n"
        "포함 가능 항목:\n"
        "- 스톡옵션 행사\n"
        "- 리스원금 상환\n"
        "- 기타재무활동",
        "Financial Model")
    for j in range(n_proj):
        _pfw(r, j, f'=0')
    r += 1

    fin_cf_row = r
    _write_row(ws, r, '재무활동현금흐름 (Financing CF)', cf['financing_cf'], years,
               bold=True, fill=LIGHT_RED)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{dividends_row}+{cl}{repurchases_row}'
             f'+{cl}{debt_iss_row}+{cl}{debt_rep_row}+{cl}{other_fin_row}',
             bold=True, fill=LIGHT_RED)
    r += 1

    _spacer(ws, r, cf_cols); r += 1

    fx_cf_row = r
    _write_row(ws, r, '외화환산효과 (FX Effect)', cf['fx_effect'], years)
    for j in range(n_proj):
        _pfw(r, j, f'=0')
    r += 1

    _spacer(ws, r, cf_cols); r += 1

    fcf_row = r
    _lbl(r, '잉여현금흐름 (Free Cash Flow = Op + Inv + Fin + FX)', bold=True)
    for i in range(n):
        _fw(r, i,
            f'={_cl(i)}{op_cf_row}+{_cl(i)}{inv_cf_row}+{_cl(i)}{fin_cf_row}+{_cl(i)}{fx_cf_row}',
            bold=True, fill=LIGHT_GREEN)
    for j in range(n_proj):
        _pfw(r, j,
             f'={_pcl(j)}{op_cf_row}+{_pcl(j)}{inv_cf_row}+{_pcl(j)}{fin_cf_row}+{_pcl(j)}{fx_cf_row}',
             bold=True, fill=LIGHT_GREEN)
    r += 1

    # FCF Margin %
    _lbl(r, '  FCF 마진율 (FCF Margin %)', ind=1)
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,{_cl(i)}{fcf_row}/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    for j in range(n_proj):
        _pfw(r, j,
             f'=IF({_pcl(j)}{rev_row}<>0,{_pcl(j)}{fcf_row}/{_pcl(j)}{rev_row},"")',
             fmt=FMT_PCT)
    r += 2

    # =========================================================================
    # CASH FLOW RECONCILIATION  (현금흐름 검증)
    # =========================================================================
    _write_section_header(
        ws, r,
        '현금흐름 검증 (Cash Flow Reconciliation: 기초현금 + 순변동 = 기말현금)',
        cols=cf_cols)
    r += 1
    _write_col_headers(ws, r, list(range(2, 2 + n)), years, start_col=2)
    _fix_ann_headers(r)
    for j, py in enumerate(proj_years):
        col = proj_start_col + j
        c = ws.cell(row=r, column=col, value=f'FY{py}E')
        _style(c, fill_hex=MED_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    r += 1

    # Beginning Cash = prior year's ending cash
    beg_cash_row = r
    _lbl(r, '기초현금 (Beginning Cash)')
    for i in range(n):
        if i == 0:
            c = ws.cell(row=r, column=2 + i, value='N/A (전기 없음)')
            _style(c, fill_hex=MED_GRAY, h_align='center', italic=True)
        elif has_ann and years[i] == ANN_YEAR:
            # Ann beginning cash = base_year cash (same as Q3 beginning)
            base_idx = years.index(_ltm_info['base_year'])
            _fw(r, i, f'={_cl(base_idx)}{cash_row}')
        else:
            prev = get_column_letter(2 + i - 1)
            _fw(r, i, f'={prev}{cash_row}')
    for j in range(n_proj):
        if j == 0:
            _pfw(r, j, f'={last_hist}{cash_row}')
        else:
            _pfw(r, j, f'={_pcl(j-1)}{beg_cash_row + 6}')
    r += 1

    # CF section totals (referenced from CF statement above)
    recon_cf_rows = []
    for label, src in [
        ('+ 영업활동현금흐름 (Operating CF)',    op_cf_row),
        ('+ 투자활동현금흐름 (Investing CF)',     inv_cf_row),
        ('+ 재무활동현금흐름 (Financing CF)',     fin_cf_row),
        ('+ 외화환산효과 (FX Effect)',            fx_cf_row),
    ]:
        recon_cf_rows.append(r)
        _lbl(r, label, ind=1)
        for i in range(n):
            _fw(r, i, f'={_cl(i)}{src}')
        for j in range(n_proj):
            _pfw(r, j, f'={_pcl(j)}{src}')
        r += 1

    # Net Change in Cash
    net_chg_row = r
    _lbl(r, '현금순변동 (Net Change in Cash)', bold=True)
    for i in range(n):
        _fw(r, i,
            f'={_cl(i)}{op_cf_row}+{_cl(i)}{inv_cf_row}+{_cl(i)}{fin_cf_row}+{_cl(i)}{fx_cf_row}',
            bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{op_cf_row}+{cl}{inv_cf_row}+{cl}{fin_cf_row}+{cl}{fx_cf_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    # Ending Cash = Beginning + Net Change
    end_cf_row = r
    assert end_cf_row == beg_cash_row + 6, \
        f"end_cf_row layout assumption broken: {end_cf_row} != {beg_cash_row + 6}"
    _lbl(r, '기말현금 (CF 기준) (Ending Cash, CF-derived)', bold=True)
    for i in range(n):
        if i == 0:
            c = ws.cell(row=r, column=2 + i, value='N/A')
            _style(c, fill_hex=MED_GRAY, h_align='center', italic=True)
        else:
            _fw(r, i, f'={_cl(i)}{beg_cash_row}+{_cl(i)}{net_chg_row}',
                bold=True, fill=LIGHT_BLUE)
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{beg_cash_row}+{_pcl(j)}{net_chg_row}',
             bold=True, fill=LIGHT_BLUE)
    r += 1

    # Ending Cash per Balance Sheet
    _lbl(r, '기말현금 (BS 기준) (Ending Cash, Balance Sheet)', bold=True)
    for i in range(n):
        _fw(r, i, f'={_cl(i)}{cash_row}', bold=True, fill=LIGHT_BLUE)
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{cash_row}', bold=True, fill=LIGHT_BLUE)
    r += 1

    # Reconciliation Check
    _lbl(r, '검증결과 (Reconciliation Check: 0 = OK | FALSE = 불일치)', bold=True)
    for i in range(n):
        if i == 0:
            c = ws.cell(row=r, column=2 + i, value='N/A')
            _style(c, fill_hex=MED_GRAY, bold=True, h_align='center')
        else:
            diff = f'ABS({_cl(i)}{end_cf_row}-{_cl(i)}{cash_row})'
            tol  = f'MAX(ABS({_cl(i)}{cash_row})*0.00001,0.5)'
            formula = f'=IF({diff}<{tol},0,FALSE)'
            c = ws.cell(row=r, column=2 + i, value=formula)
            _style(c, fill_hex=YELLOW, bold=True, h_align='center')
    for j in range(n_proj):
        c = ws.cell(row=r, column=proj_start_col + j, value=0)
        _style(c, fill_hex=YELLOW, bold=True, h_align='center',
               number_format=FMT_KRW)
    r += 2

    # =========================================================================
    # PP&E SCHEDULE  (유형자산 증감표)
    # =========================================================================
    _write_section_header(ws, r, '유형자산 증감표 (PP&E Schedule)', cols=total_cols); r += 1
    _write_col_headers(ws, r, list(range(2, 2 + n)), years, start_col=2)
    _fix_ann_headers(r)
    for j, py in enumerate(proj_years):
        col = proj_start_col + j
        c = ws.cell(row=r, column=col, value=f'FY{py}E')
        _style(c, fill_hex=MED_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    # Step-up column header
    step_col = total_cols + 1
    step_col_letter = get_column_letter(step_col)
    c = ws.cell(row=r, column=step_col, value='Step-up')
    _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE, h_align='center')
    c.border = THIN_BOX
    ws.column_dimensions[step_col_letter].width = 14
    r += 1

    # Pre-compute row positions
    ppe_beg_sch_row    = r
    ppe_capex_sch_row  = r + 1
    ppe_depr_sch_row   = r + 2
    ppe_end_sch_row    = r + 3
    # r + 4 = spacer
    da_capex_pct_row   = r + 5
    capex_pct_rev_row  = r + 6

    step_up_ref   = f'${step_col_letter}${da_capex_pct_row}'
    capex_pct_ref = f'${step_col_letter}${capex_pct_rev_row}'

    # --- Beginning PP&E ---
    _lbl(r, '기초 유형자산 (Beginning PP&E)')
    for i in range(n):
        if i == 0:
            c = ws.cell(row=r, column=2 + i, value='N/A')
            _style(c, fill_hex=MED_GRAY, h_align='center', italic=True)
        elif has_ann and years[i] == ANN_YEAR:
            # Ann beginning = same as Q3 beginning = base_year PP&E
            base_idx = years.index(_ltm_info['base_year'])
            _fw(r, i, f'={_cl(base_idx)}{ppe_row}')
        else:
            _fw(r, i, f'={_cl(i-1)}{ppe_row}')
    for j in range(n_proj):
        if j == 0:
            _pfw(r, j, f'={last_hist}{ppe_row}')
        else:
            _pfw(r, j, f'={_pcl(j-1)}{ppe_end_sch_row}')
    r += 1

    # --- + Capital Expenditures ---
    _lbl(r, '+ 유형자산취득 (+ Capex)')
    for i in range(n):
        _fw(r, i, f'=ABS({_cl(i)}{capex_row})')
    _capex_ref_yr = ANN_YEAR if has_ann else latest_yr
    last_hist_capex_val = abs(_val(cf['capex'], _capex_ref_yr) or 0)
    for j in range(n_proj):
        if j < 2:
            c = ws.cell(row=r, column=proj_start_col + j,
                        value=round(last_hist_capex_val, 1))
            _style(c, fill_hex=YELLOW, h_align='right', number_format=FMT_KRW)
            c.border = THIN_BOX
        else:
            _pfw(r, j, f'={_pcl(j)}{rev_row}*{capex_pct_ref}')
    r += 1

    # --- - Depreciation ---
    _lbl(r, '- 감가상각비 (- Depreciation)')
    for i in range(n):
        _fw(r, i, f'={_cl(i)}{da_row}')
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{ppe_capex_sch_row}*{_pcl(j)}{da_capex_pct_row}')
    r += 1

    # --- Ending PP&E = Beg + Capex - Depr ---
    _lbl(r, '기말 유형자산 (Ending PP&E)', bold=True)
    for i in range(n):
        _fw(r, i, f'={_cl(i)}{ppe_row}', bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{ppe_beg_sch_row}+{cl}{ppe_capex_sch_row}-{cl}{ppe_depr_sch_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 1

    _spacer(ws, r, total_cols); r += 1

    # --- D&A / Capex % ---
    _lbl(r, '감가상각비 / Capex 비율 (D&A / Capex %)')
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{ppe_capex_sch_row}<>0,'
            f'{_cl(i)}{ppe_depr_sch_row}/{_cl(i)}{ppe_capex_sch_row},"")',
            fmt=FMT_PCT)
    avg_start = max(0, n - 3)
    avg_cells = ','.join(f'{_cl(i)}{da_capex_pct_row}' for i in range(avg_start, n))
    for j in range(n_proj):
        if j == 0:
            _pfw(r, j, f'=AVERAGE({avg_cells})', fmt=FMT_PCT)
        else:
            _pfw(r, j, f'={_pcl(j-1)}{da_capex_pct_row}+{step_up_ref}', fmt=FMT_PCT)
    # Step-up editable cell (default 2.0%)
    c = ws.cell(row=r, column=step_col, value=0.02)
    _style(c, fill_hex=YELLOW, h_align='right', number_format=FMT_PCT)
    c.border = THIN_BOX
    r += 1

    # --- Capex % of Revenue (for projection years 3-5) ---
    _lbl(r, 'Capex / 매출액 비율 (Capex % of Revenue, Yr 3-5)')
    for i in range(n):
        _fw(r, i,
            f'=IF({_cl(i)}{rev_row}<>0,'
            f'ABS({_cl(i)}{capex_row})/{_cl(i)}{rev_row},"")',
            fmt=FMT_PCT)
    capex_rev_vals = []
    for yr in years:
        cap_v = abs((_val(cf['capex'], yr) or 0))
        rev_v = _val(inc['revenue'], yr)
        if cap_v > 0 and rev_v and rev_v > 0:
            capex_rev_vals.append(cap_v / rev_v)
    avg_capex_pct_rev = _safe_avg(capex_rev_vals) if capex_rev_vals else 0.05
    c = ws.cell(row=r, column=step_col, value=round(avg_capex_pct_rev, 4))
    _style(c, fill_hex=YELLOW, h_align='right', number_format=FMT_PCT)
    c.border = THIN_BOX
    r += 2

    # =========================================================================
    # RETAINED EARNINGS SCHEDULE  (이익잉여금 증감표)
    # =========================================================================
    _write_section_header(ws, r, '이익잉여금 증감표 (Retained Earnings Schedule)', cols=total_cols); r += 1
    _write_col_headers(ws, r, list(range(2, 2 + n)), years, start_col=2)
    _fix_ann_headers(r)
    for j, py in enumerate(proj_years):
        col = proj_start_col + j
        c = ws.cell(row=r, column=col, value=f'FY{py}E')
        _style(c, fill_hex=MED_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    r += 1

    # Pre-compute row positions
    re_beg_sch_row  = r
    re_ni_sch_row   = r + 1
    re_div_sch_row  = r + 2
    re_rep_sch_row  = r + 3
    re_end_sch_row  = r + 4

    # --- Beginning Retained Earnings ---
    _lbl(r, '기초 이익잉여금 (Beginning RE)')
    for i in range(n):
        if i == 0:
            c = ws.cell(row=r, column=2 + i, value='N/A')
            _style(c, fill_hex=MED_GRAY, h_align='center', italic=True)
        elif has_ann and years[i] == ANN_YEAR:
            base_idx = years.index(_ltm_info['base_year'])
            _fw(r, i, f'={_cl(base_idx)}{re_row}')
        else:
            _fw(r, i, f'={_cl(i-1)}{re_row}')
    for j in range(n_proj):
        if j == 0:
            _pfw(r, j, f'={last_hist}{re_row}')
        else:
            _pfw(r, j, f'={_pcl(j-1)}{re_end_sch_row}')
    r += 1

    # --- + Net Income ---
    _lbl(r, '+ 당기순이익 (+ Net Income)')
    for i in range(n):
        _fw(r, i, f'={_cl(i)}{ni_row}')
    for j in range(n_proj):
        _pfw(r, j, f'={_pcl(j)}{ni_row}')
    r += 1

    # --- - Dividends ---
    _lbl(r, '- 배당금 (- Dividends)')
    for i in range(n):
        _fw(r, i, f'=ABS({_cl(i)}{dividends_row})')
    for j in range(n_proj):
        _pfw(r, j, f'=ABS({_pcl(j)}{dividends_row})')
    r += 1

    # --- - Share Repurchases ---
    _lbl(r, '- 자사주매입 (- Repurchases)')
    for i in range(n):
        _fw(r, i, f'=ABS({_cl(i)}{repurchases_row})')
    for j in range(n_proj):
        prev = _cl(n - 1) if j == 0 else _pcl(j - 1)
        _pfw(r, j, f'=ABS({prev}{repurchases_row})')
    r += 1

    # --- Ending Retained Earnings = Beg + NI - Div - Repurchases ---
    _lbl(r, '기말 이익잉여금 (Ending RE)', bold=True)
    for i in range(n):
        _fw(r, i, f'={_cl(i)}{re_row}', bold=True, fill=XLIGHT_BLUE)
    for j in range(n_proj):
        cl = _pcl(j)
        _pfw(r, j,
             f'={cl}{re_beg_sch_row}+{cl}{re_ni_sch_row}'
             f'-{cl}{re_div_sch_row}-{cl}{re_rep_sch_row}',
             bold=True, fill=XLIGHT_BLUE)
    r += 2

    # ── Backfill: BS Retained Earnings -> RE schedule ending balance ──
    for j in range(n_proj):
        _pfw(re_row, j, f'={_pcl(j)}{re_end_sch_row}')

    # ── Backfill: CF Share Repurchases -> negated RE schedule repurchases ─
    for j in range(n_proj):
        _pfw(repurchases_row, j, f'=-{_pcl(j)}{re_rep_sch_row}')

    # ── Backfill: BS PP&E projections -> PP&E schedule ending balance ──
    for j in range(n_proj):
        _pfw(ppe_row, j, f'={_pcl(j)}{ppe_end_sch_row}')

    # ── Backfill: IS D&A projections -> PP&E schedule depreciation ─────
    for j in range(n_proj):
        _pfw(da_row, j, f'={_pcl(j)}{ppe_depr_sch_row}')

    # ── Backfill: CF D&A projections -> PP&E schedule depreciation ─────
    for j in range(n_proj):
        _pfw(cf_da_row, j, f'={_pcl(j)}{ppe_depr_sch_row}')

    # ── Backfill: CF Capex projections -> PP&E schedule capex ──────────
    for j in range(n_proj):
        _pfw(capex_row, j, f'=-{_pcl(j)}{ppe_capex_sch_row}')

    # ── Backfill: BS Cash projections -> CF Reconciliation ending cash ─
    for j in range(n_proj):
        _pfw(cash_row, j, f'={_pcl(j)}{end_cf_row}')

    ws.freeze_panes = 'A3'

    # =========================================================================
    # REVENUE CAGR  (매출액 연평균성장률)
    # =========================================================================
    _latest_hist_cl = get_column_letter(2 + n - 1)
    # Find earliest year with actual revenue data for CAGR start.
    # Fall back progressively until we find a usable starting point.
    # Use only actual annual years (exclude Q3 cum / annualized)
    _annual_yrs = [yr for yr in years if isinstance(yr, int) and
                   (not has_ann or yr != _ltm_info.get('ltm_year'))]
    _cagr_start_year = _annual_yrs[0] if _annual_yrs else years[0]
    _cagr_start_rev = 0
    for yr in (_annual_yrs or years):
        rev_val = inc['revenue'].get(yr)
        if rev_val and rev_val > 0:
            _cagr_start_year = yr
            _cagr_start_rev = rev_val
            break
    # Periods = gap between start year and latest int year
    _cagr_periods = latest_yr - _cagr_start_year
    if _cagr_periods < 1:
        _cagr_periods = 1
    _cagr_end_label = (_ltm_info['ann_label'] if has_ann
                       else f'FY{latest_yr}')

    r += 2
    _write_section_header(ws, r, '매출액 CAGR (Revenue CAGR)', cols=total_cols); r += 1

    _lbl(r, f'기초 매출액 (Starting Revenue, FY{_cagr_start_year})')
    c = ws.cell(row=r, column=2, value=round(_cagr_start_rev, 1))
    _style(c, fill_hex=YELLOW, bold=False, h_align='right',
           number_format=FMT_KRW)
    c.border = THIN_BOX
    cagr_start_row = r; r += 1

    _lbl(r, f'기말 매출액 (Ending Revenue, {_cagr_end_label})')
    c = ws.cell(row=r, column=2, value=f'={_latest_hist_cl}{rev_row}')
    _style(c, fill_hex=LIGHT_GREEN, bold=False, h_align='right',
           number_format=FMT_KRW)
    c.border = THIN_BOX
    cagr_end_row = r; r += 1

    _lbl(r, '기간 (Number of Periods)')
    c = ws.cell(row=r, column=2, value=_cagr_periods)
    _style(c, fill_hex=YELLOW, bold=False, h_align='right',
           number_format='0')
    c.border = THIN_BOX
    cagr_periods_row = r; r += 1

    _lbl(r, '매출액 CAGR (Revenue CAGR)', bold=True)
    c = ws.cell(row=r, column=2,
                value=f'=IF(B{cagr_start_row}>0,(B{cagr_end_row}/B{cagr_start_row})^(1/B{cagr_periods_row})-1,0)')
    _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right',
           number_format=FMT_PCT)
    c.border = THIN_BOX
    cagr_row = r; r += 1

    # Return row map so DCF sheet can build cross-sheet references.
    latest_col = get_column_letter(2 + n - 1)
    fs_rows = {
        'latest_col':         latest_col,
        # Income Statement
        'revenue':            rev_row,
        'cogs':               cogs_row,
        'gp':                 gp_row,
        'operating_income':   oi_row,
        'da':                 da_row,
        'ebitda':             ebitda_row,
        'interest_expense':   int_exp_row,
        'interest_income':    int_inc_row,
        'other_income':       other_inc_row,
        'pretax_income':      pretax_row,
        'tax_expense':        tax_row,
        'net_income':         ni_row,
        'eps_diluted':        eps_diluted_row,
        'shares_diluted':     shares_diluted_row,
        # Assumptions band
        'rev_growth_asm':     rev_growth_asm_row,
        'gp_margin_asm':      gp_margin_asm_row,
        'sga_pct_asm':        sga_pct_asm_row,
        'oi_margin_asm':      oi_margin_asm_row,
        'tax_rate_asm':       tax_rate_asm_row,
        'da_pct_asm':         da_pct_asm_row,
        # Projection metadata
        'proj_start_col':     proj_start_col,
        'n_proj':             n_proj,
        'n_hist':             n,
        'total_cols':         total_cols,
        'dropdown_ref':       dropdown_ref,
        'scenario_rows':      scenario_rows,
    }
    # Add additional FS row references for validation and DCF
    fs_rows.update({
        # Balance Sheet
        'cash':               cash_row,
        'st_investments':     st_inv_row,
        'accounts_rec':       ar_row,
        'inventory':          inventory_row,
        'other_current_a':    other_ca_row,
        'total_current_a':    total_ca_row,
        'ppe':                ppe_row,
        'goodwill':           gw_row,
        'intangibles':        intangibles_row,
        'lt_investments':     lt_inv_row,
        'other_noncurrent_a': other_nca_row,
        'total_assets':       total_assets_row,
        'accounts_pay':       ap_row,
        'accrued_liab':       accrued_row,
        'st_debt':            st_debt_row,
        'deferred_rev_cur':   deferred_rev_row,
        'other_current_l':    other_cl_row,
        'total_current_l':    total_cl_row,
        'lt_debt':            lt_debt_row,
        'deferred_tax_l':     dtl_row,
        'other_noncurrent_l': other_ncl_row,
        'total_liabilities':  total_liabilities_row,
        'common_stock':       cs_row,
        'apic':               apic_row,
        'retained_earnings':  re_row,
        'treasury_stock':     ts_row,
        'total_equity':       total_equity_row,
        # Cash Flow
        'cf_net_income':      cf_ni_row,
        'cf_da':              cf_da_row,
        'operating_cf':       r,  # placeholder -- actual op_cf_row
        'capex':              capex_row,
        'investing_cf':       inv_cf_row,
        'dividends':          dividends_row,
        'repurchases':        repurchases_row,
        'financing_cf':       fin_cf_row,
        'fx_effect':          fx_cf_row,
        'fcf':                fcf_row,
        # PP&E schedule
        'ppe_capex_sch':      ppe_capex_sch_row,
        'ppe_depr_sch':       ppe_depr_sch_row,
        'ppe_end_sch':        ppe_end_sch_row,
    })
    # Fix operating_cf reference
    fs_rows['operating_cf'] = op_cf_row

    return fs_rows


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 -- WACC  (가중평균자본비용)
# ─────────────────────────────────────────────────────────────────────────────

def _write_wacc_sheet(ws, company_info, financial_data, fs_rows):
    """Write the WACC sheet.  Returns wacc_rows dict for DCF cross-references."""

    def _lbl(row_num, text, bold=False, fill=None, ind=0):
        c = ws.cell(row=row_num, column=1, value='  ' * ind + text)
        _style(c, fill_hex=fill, bold=bold)
        return c

    wacc_inputs = financial_data.get('wacc_inputs', {})
    comparables = wacc_inputs.get('comparables', [])
    shares_bk = wacc_inputs.get('shares_breakdown', {})
    current_price_data = wacc_inputs.get('current_price', {})
    price = current_price_data.get('price') or 0
    price_date = current_price_data.get('date', '')

    corp_name = company_info.get('corp_name', '')
    stock_code = company_info.get('stock_code', '')
    fs_sheet = "'Financial Statements'"

    _set_col_widths(ws, {1: 38, 2: 16, 3: 16, 4: 16, 5: 14, 6: 14, 7: 14, 8: 14})

    r = 1
    # Title
    title = ws.cell(row=r, column=1,
                    value=f"{corp_name} ({stock_code}) - 가중평균자본비용 (WACC)")
    _style(title, fill_hex=DARK_BLUE, bold=True, font_color=WHITE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1

    sub = ws.cell(row=r, column=1,
                  value="노란색 셀 = 사용자 입력  |  녹색 셀 = 수식  |  파란색 셀 = 참조")
    _style(sub, fill_hex=XLIGHT_BLUE, italic=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 2

    # ── Share Price & Market Cap ─────────────────────────────────────────
    _write_section_header(ws, r, '주가 및 시가총액 (Share Price & Market Cap)', cols=3)
    r += 1

    price_row = r
    _lbl(r, f'현재 주가 (Share Price, {price_date})')
    c = ws.cell(row=r, column=2, value=price)
    _style(c, fill_hex=YELLOW, bold=False, h_align='right', number_format='#,##0')
    c.border = THIN_BOX
    r += 1

    # Shares breakdown
    _lbl(r, '보통주 (Basic Shares)')
    c = ws.cell(row=r, column=2, value=shares_bk.get('basic', 0))
    _style(c, fill_hex=YELLOW, h_align='right', number_format=FMT_SHARES)
    c.border = THIN_BOX
    basic_row = r; r += 1

    _lbl(r, '+ RSU / 주식보상 (RSUs)')
    c = ws.cell(row=r, column=2, value=shares_bk.get('rsus', 0))
    _style(c, fill_hex=YELLOW, h_align='right', number_format=FMT_SHARES)
    c.border = THIN_BOX
    rsu_row = r; r += 1

    _lbl(r, '+ 스톡옵션 (Stock Options)')
    c = ws.cell(row=r, column=2, value=shares_bk.get('options', 0))
    _style(c, fill_hex=YELLOW, h_align='right', number_format=FMT_SHARES)
    c.border = THIN_BOX
    opt_row = r; r += 1

    _lbl(r, '+ 전환사채 (Convertible Debt)')
    c = ws.cell(row=r, column=2, value=shares_bk.get('conv_debt', 0))
    _style(c, fill_hex=YELLOW, h_align='right', number_format=FMT_SHARES)
    c.border = THIN_BOX
    cd_row = r; r += 1

    _lbl(r, '+ 전환우선주 (Convertible Preferred)')
    c = ws.cell(row=r, column=2, value=shares_bk.get('conv_pref', 0))
    _style(c, fill_hex=YELLOW, h_align='right', number_format=FMT_SHARES)
    c.border = THIN_BOX
    cp_row = r; r += 1

    diluted_shares_row = r
    _lbl(r, '희석주식수 (Diluted Shares)', bold=True)
    c = ws.cell(row=r, column=2,
                value=f'=B{basic_row}+B{rsu_row}+B{opt_row}+B{cd_row}+B{cp_row}')
    _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right',
           number_format=FMT_SHARES)
    c.border = THIN_BOX
    r += 1

    mktcap_row = r
    _lbl(r, '시가총액 (Market Cap)', bold=True)
    c = ws.cell(row=r, column=2, value=f'=B{price_row}*B{diluted_shares_row}')
    _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right',
           number_format='#,##0')
    c.border = THIN_BOX
    r += 2

    # ── Cost of Debt ─────────────────────────────────────────────────────
    _write_section_header(ws, r, '타인자본비용 (Cost of Debt)', cols=3)
    r += 1

    cod_row = r
    implied_cod = wacc_inputs.get('implied_cod', 0.05)
    _lbl(r, '내재 차입비용 (Implied Cost of Debt)')
    c = ws.cell(row=r, column=2, value=implied_cod)
    _style(c, fill_hex=YELLOW, h_align='right', number_format=FMT_PCT)
    c.border = THIN_BOX
    r += 1

    # Tax rate from FS
    latest_col = fs_rows.get('latest_col', 'B')
    tax_rate_row = r
    _lbl(r, '유효법인세율 (Effective Tax Rate)')
    c = ws.cell(row=r, column=2,
                value=f'=IF({fs_sheet}!{latest_col}{fs_rows["pretax_income"]}>0,'
                      f'{fs_sheet}!{latest_col}{fs_rows["tax_expense"]}'
                      f'/{fs_sheet}!{latest_col}{fs_rows["pretax_income"]},0.22)')
    _style(c, fill_hex=LIGHT_BLUE, h_align='right', number_format=FMT_PCT)
    c.border = THIN_BOX
    r += 1

    cod_at_row = r
    _lbl(r, '세후 타인자본비용 (After-tax Cost of Debt)', bold=True)
    c = ws.cell(row=r, column=2, value=f'=B{cod_row}*(1-B{tax_rate_row})')
    _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right',
           number_format=FMT_PCT)
    c.border = THIN_BOX
    r += 2

    # ── Cost of Equity (CAPM) ────────────────────────────────────────────
    _write_section_header(ws, r, '자기자본비용 (Cost of Equity - CAPM)', cols=3)
    r += 1

    rf_row = r
    rf_rate = wacc_inputs.get('treasury_yield', 0.035)
    _lbl(r, '무위험이자율 (Risk-Free Rate, KR 10Y)')
    c = ws.cell(row=r, column=2, value=rf_rate)
    _style(c, fill_hex=YELLOW, h_align='right', number_format=FMT_PCT)
    c.border = THIN_BOX
    r += 1

    erp_row = r
    erp = wacc_inputs.get('kroll_erp', 0.065)
    _lbl(r, '주식위험프리미엄 (Equity Risk Premium)')
    c = ws.cell(row=r, column=2, value=erp)
    _style(c, fill_hex=YELLOW, h_align='right', number_format=FMT_PCT)
    c.border = THIN_BOX
    r += 2

    # ── Comparable Company Betas ─────────────────────────────────────────
    _write_section_header(ws, r,
        '비교기업 베타 (Comparable Company Betas)', cols=8)
    r += 1

    # Column headers for comps table
    comp_headers = ['기업명', '레버드 베타', '주가', '주식수', '시가총액',
                    '순부채', '법인세율', '언레버드 베타']
    for ci, h in enumerate(comp_headers, 1):
        c = ws.cell(row=r, column=ci, value=h)
        _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    r += 1

    comp_start_row = r
    for idx, comp in enumerate(comparables[:10]):
        cr = r + idx
        ws.cell(row=cr, column=1, value=comp.get('name', '')).border = THIN_BOX
        c = ws.cell(row=cr, column=2, value=comp.get('beta', 1.0))
        _style(c, fill_hex=YELLOW, h_align='right', number_format='0.00')
        c.border = THIN_BOX

        c = ws.cell(row=cr, column=3, value=comp.get('price', 0))
        _style(c, h_align='right', number_format='#,##0')
        c.border = THIN_BOX

        shares_val = comp.get('shares', 0)
        c = ws.cell(row=cr, column=4, value=shares_val)
        _style(c, h_align='right', number_format=FMT_SHARES)
        c.border = THIN_BOX

        # Market Cap = Price * Shares
        c = ws.cell(row=cr, column=5, value=f'=C{cr}*D{cr}')
        _style(c, fill_hex=LIGHT_GREEN, h_align='right', number_format='#,##0')
        c.border = THIN_BOX

        net_debt = comp.get('net_debt', 0)
        c = ws.cell(row=cr, column=6, value=net_debt)
        _style(c, fill_hex=YELLOW, h_align='right', number_format='#,##0')
        c.border = THIN_BOX

        tax_r = comp.get('tax_rate', 0.22)
        c = ws.cell(row=cr, column=7, value=tax_r)
        _style(c, fill_hex=YELLOW, h_align='right', number_format=FMT_PCT)
        c.border = THIN_BOX

        # Unlevered Beta = Levered Beta / [1 + (1-Tax)*(NetDebt/MktCap)]
        c = ws.cell(row=cr, column=8,
                    value=f'=IF(E{cr}>0,B{cr}/(1+(1-G{cr})*(F{cr}/E{cr})),B{cr})')
        _style(c, fill_hex=LIGHT_GREEN, h_align='right', number_format='0.00')
        c.border = THIN_BOX

    n_comps = min(len(comparables), 10)
    comp_end_row = comp_start_row + max(n_comps - 1, 0)

    # Fill empty rows if fewer than 10 comps
    for idx in range(n_comps, 10):
        cr = comp_start_row + idx
        for ci in range(1, 9):
            ws.cell(row=cr, column=ci).border = THIN_BOX

    r = comp_start_row + 10
    r += 1

    # ── Average Unlevered Beta & Re-levering ──────────────────────────
    avg_ul_beta_row = r
    _lbl(r, '평균 언레버드 베타 (Avg Unlevered Beta)', bold=True)
    if n_comps > 0:
        avg_range = f'H{comp_start_row}:H{comp_end_row}'
        c = ws.cell(row=r, column=2,
                    value=f'=AVERAGE({avg_range})')
    else:
        c = ws.cell(row=r, column=2, value=1.0)
    _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right',
           number_format='0.00')
    c.border = THIN_BOX
    r += 2

    # Target company re-levering
    _write_section_header(ws, r,
        '대상기업 리레버링 (Re-levering to Target)', cols=3)
    r += 1

    tgt_net_debt_row = r
    _lbl(r, '순부채 (Net Debt = 단기+장기차입금-현금)')
    c = ws.cell(row=r, column=2,
                value=f'={fs_sheet}!{latest_col}{fs_rows["st_debt"]}'
                      f'+{fs_sheet}!{latest_col}{fs_rows["lt_debt"]}'
                      f'-{fs_sheet}!{latest_col}{fs_rows["cash"]}')
    _style(c, fill_hex=LIGHT_BLUE, h_align='right', number_format='#,##0')
    c.border = THIN_BOX
    r += 1

    tgt_mktcap_row = r
    _lbl(r, '시가총액 (Market Cap)')
    c = ws.cell(row=r, column=2, value=f'=B{mktcap_row}')
    _style(c, fill_hex=LIGHT_BLUE, h_align='right', number_format='#,##0')
    c.border = THIN_BOX
    r += 1

    relevered_beta_row = r
    _lbl(r, '리레버드 베타 (Re-levered Beta)', bold=True)
    c = ws.cell(row=r, column=2,
                value=f'=IF(B{tgt_mktcap_row}>0,'
                      f'B{avg_ul_beta_row}*(1+(1-B{tax_rate_row})'
                      f'*(B{tgt_net_debt_row}/B{tgt_mktcap_row})),'
                      f'B{avg_ul_beta_row})')
    _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right',
           number_format='0.00')
    c.border = THIN_BOX
    r += 1

    coe_row = r
    _lbl(r, '자기자본비용 (Cost of Equity = Rf + β × ERP)', bold=True)
    c = ws.cell(row=r, column=2,
                value=f'=B{rf_row}+B{relevered_beta_row}*B{erp_row}')
    _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right',
           number_format=FMT_PCT)
    c.border = THIN_BOX
    r += 2

    # ── Capital Structure & WACC ─────────────────────────────────────────
    _write_section_header(ws, r,
        '자본구조 및 WACC (Capital Structure & WACC)', cols=3)
    r += 1

    ev_row = r
    _lbl(r, '기업가치 (Enterprise Value = MktCap + NetDebt)')
    c = ws.cell(row=r, column=2,
                value=f'=B{tgt_mktcap_row}+B{tgt_net_debt_row}')
    _style(c, fill_hex=LIGHT_GREEN, h_align='right', number_format='#,##0')
    c.border = THIN_BOX
    r += 1

    eq_weight_row = r
    _lbl(r, '자기자본 비중 (Equity Weight)')
    c = ws.cell(row=r, column=2,
                value=f'=IF(B{ev_row}>0,B{tgt_mktcap_row}/B{ev_row},1)')
    _style(c, fill_hex=LIGHT_GREEN, h_align='right', number_format=FMT_PCT)
    c.border = THIN_BOX
    r += 1

    debt_weight_row = r
    _lbl(r, '타인자본 비중 (Debt Weight)')
    c = ws.cell(row=r, column=2,
                value=f'=IF(B{ev_row}>0,MAX(B{tgt_net_debt_row},0)/B{ev_row},0)')
    _style(c, fill_hex=LIGHT_GREEN, h_align='right', number_format=FMT_PCT)
    c.border = THIN_BOX
    r += 2

    wacc_row = r
    _lbl(r, 'WACC (가중평균자본비용)', bold=True)
    c = ws.cell(row=r, column=2,
                value=f'=B{eq_weight_row}*B{coe_row}+B{debt_weight_row}*B{cod_at_row}')
    _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right',
           number_format=FMT_PCT)
    c.border = BOT_THICK
    c.font = _font(bold=True, size=12)

    wacc_rows = {
        'price':           price_row,
        'diluted_shares':  diluted_shares_row,
        'mktcap':          mktcap_row,
        'cod':             cod_row,
        'tax_rate':        tax_rate_row,
        'cod_at':          cod_at_row,
        'rf_rate':         rf_row,
        'erp':             erp_row,
        'relevered_beta':  relevered_beta_row,
        'coe':             coe_row,
        'net_debt':        tgt_net_debt_row,
        'ev':              ev_row,
        'eq_weight':       eq_weight_row,
        'debt_weight':     debt_weight_row,
        'wacc':            wacc_row,
    }
    return wacc_rows


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3 -- DCF MODEL  (할인현금흐름 모형)
# ─────────────────────────────────────────────────────────────────────────────

def _write_dcf_model(ws, company_info, financial_data, fs_rows, wacc_rows):
    """Write the DCF Model sheet."""

    def _lbl(row_num, text, bold=False, fill=None, ind=0):
        c = ws.cell(row=row_num, column=1, value='  ' * ind + text)
        _style(c, fill_hex=fill, bold=bold)
        return c

    years = financial_data['years']
    inc = financial_data['income_statement']
    cf = financial_data['cash_flow']
    latest_yr = sorted(years)[-1]

    corp_name = company_info.get('corp_name', '')
    stock_code = company_info.get('stock_code', '')
    fs_sheet = "'Financial Statements'"
    wacc_sheet = "'WACC'"

    n_hist = fs_rows.get('n_hist', len(years))
    proj_start_col = fs_rows.get('proj_start_col', 2 + n_hist)
    n_proj = fs_rows.get('n_proj', 5)
    proj_years = [latest_yr + i for i in range(1, n_proj + 1)]
    latest_col = fs_rows.get('latest_col', get_column_letter(2 + n_hist - 1))

    _set_col_widths(ws, {1: 40, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16, 8: 16})

    def _pcl(j):
        return get_column_letter(proj_start_col + j)

    r = 1
    # Title
    title = ws.cell(row=r, column=1,
                    value=f"{corp_name} ({stock_code})"
                          f" - 할인현금흐름 모형 (DCF Model)  |  단위: KRW")
    _style(title, fill_hex=DARK_BLUE, bold=True, font_color=WHITE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 2

    # ── Key Assumptions ──────────────────────────────────────────────────
    _write_section_header(ws, r, '주요 가정 (Key Assumptions)', cols=3)
    r += 1

    wacc_ref_row = r
    _lbl(r, 'WACC (가중평균자본비용)')
    c = ws.cell(row=r, column=2, value=f'={wacc_sheet}!B{wacc_rows["wacc"]}')
    _style(c, fill_hex=LIGHT_BLUE, bold=True, h_align='right',
           number_format=FMT_PCT)
    c.border = THIN_BOX
    r += 1

    tax_ref_row = r
    _lbl(r, '유효법인세율 (Tax Rate)')
    c = ws.cell(row=r, column=2, value=f'={wacc_sheet}!B{wacc_rows["tax_rate"]}')
    _style(c, fill_hex=LIGHT_BLUE, h_align='right', number_format=FMT_PCT)
    c.border = THIN_BOX
    r += 1

    tg_row = r
    _lbl(r, '영구성장률 (Terminal Growth Rate)')
    c = ws.cell(row=r, column=2, value=0.025)
    _style(c, fill_hex=YELLOW, h_align='right', number_format=FMT_PCT)
    c.border = THIN_BOX
    r += 2

    # ── DCF Projection Table ─────────────────────────────────────────────
    _write_section_header(ws, r, 'DCF 추정표 (DCF Projection)', cols=8)
    r += 1

    # Column headers: Base Year + 5 projection years
    c = ws.cell(row=r, column=1, value='')
    _style(c, fill_hex=DARK_BLUE)
    c = ws.cell(row=r, column=2, value=f'FY{latest_yr}')
    _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE, h_align='center')
    c.border = THIN_BOX
    for j, py in enumerate(proj_years):
        col = 3 + j
        c = ws.cell(row=r, column=col, value=f'FY{py}E')
        _style(c, fill_hex=MED_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    r += 1

    def _dcf_col(j):
        return get_column_letter(3 + j)

    def _dcf_prev(j):
        return 'B' if j == 0 else _dcf_col(j - 1)

    # Revenue
    dcf_rev_row = r
    _lbl(r, '매출액 (Revenue)')
    c = ws.cell(row=r, column=2,
                value=f'={fs_sheet}!{latest_col}{fs_rows["revenue"]}')
    _style(c, h_align='right', number_format=FMT_KRW)
    for j in range(n_proj):
        c = ws.cell(row=r, column=3 + j,
                    value=f'={fs_sheet}!{_pcl(j)}{fs_rows["revenue"]}')
        _style(c, fill_hex=XLIGHT_BLUE if j % 2 == 0 else WHITE,
               h_align='right', number_format=FMT_KRW)
    r += 1

    # Revenue Growth
    dcf_rev_g_row = r
    _lbl(r, '  매출 성장률 (Revenue Growth %)', ind=1)
    c = ws.cell(row=r, column=2, value='')
    _style(c, fill_hex=MED_GRAY, h_align='center')
    for j in range(n_proj):
        c = ws.cell(row=r, column=3 + j,
                    value=f'=IF({_dcf_prev(j)}{dcf_rev_row}<>0,'
                          f'{_dcf_col(j)}{dcf_rev_row}/{_dcf_prev(j)}{dcf_rev_row}-1,"")')
        _style(c, h_align='right', number_format=FMT_PCT)
    r += 1

    # EBITDA
    dcf_ebitda_row = r
    _lbl(r, 'EBITDA')
    c = ws.cell(row=r, column=2,
                value=f'={fs_sheet}!{latest_col}{fs_rows["ebitda"]}')
    _style(c, h_align='right', number_format=FMT_KRW)
    for j in range(n_proj):
        c = ws.cell(row=r, column=3 + j,
                    value=f'={fs_sheet}!{_pcl(j)}{fs_rows["ebitda"]}')
        _style(c, fill_hex=XLIGHT_BLUE if j % 2 == 0 else WHITE,
               h_align='right', number_format=FMT_KRW)
    r += 1

    # D&A
    dcf_da_row = r
    _lbl(r, '  감가상각비 (D&A)', ind=1)
    c = ws.cell(row=r, column=2,
                value=f'={fs_sheet}!{latest_col}{fs_rows["da"]}')
    _style(c, h_align='right', number_format=FMT_KRW)
    for j in range(n_proj):
        c = ws.cell(row=r, column=3 + j,
                    value=f'={fs_sheet}!{_pcl(j)}{fs_rows["da"]}')
        _style(c, h_align='right', number_format=FMT_KRW)
    r += 1

    # EBIT = EBITDA - D&A
    dcf_ebit_row = r
    _lbl(r, '영업이익 (EBIT = EBITDA - D&A)', bold=True)
    c = ws.cell(row=r, column=2,
                value=f'=B{dcf_ebitda_row}-B{dcf_da_row}')
    _style(c, bold=True, h_align='right', number_format=FMT_KRW, fill_hex=XLIGHT_BLUE)
    for j in range(n_proj):
        cl = _dcf_col(j)
        c = ws.cell(row=r, column=3 + j,
                    value=f'={cl}{dcf_ebitda_row}-{cl}{dcf_da_row}')
        _style(c, bold=True, fill_hex=XLIGHT_BLUE, h_align='right',
               number_format=FMT_KRW)
    r += 1

    # NOPAT = EBIT * (1 - Tax Rate)
    dcf_nopat_row = r
    _lbl(r, 'NOPAT (세후영업이익)', bold=True)
    c = ws.cell(row=r, column=2,
                value=f'=B{dcf_ebit_row}*(1-$B${tax_ref_row})')
    _style(c, bold=True, h_align='right', number_format=FMT_KRW)
    for j in range(n_proj):
        cl = _dcf_col(j)
        c = ws.cell(row=r, column=3 + j,
                    value=f'={cl}{dcf_ebit_row}*(1-$B${tax_ref_row})')
        _style(c, bold=True, h_align='right', number_format=FMT_KRW)
    r += 1

    # + D&A
    dcf_add_da_row = r
    _lbl(r, '+ 감가상각비 (+ D&A)', ind=1)
    c = ws.cell(row=r, column=2, value=f'=B{dcf_da_row}')
    _style(c, h_align='right', number_format=FMT_KRW)
    for j in range(n_proj):
        cl = _dcf_col(j)
        c = ws.cell(row=r, column=3 + j, value=f'={cl}{dcf_da_row}')
        _style(c, h_align='right', number_format=FMT_KRW)
    r += 1

    # - Capex
    dcf_capex_row = r
    _lbl(r, '- 자본적지출 (- Capex)', ind=1)
    c = ws.cell(row=r, column=2,
                value=f'=ABS({fs_sheet}!{latest_col}{fs_rows["capex"]})')
    _style(c, h_align='right', number_format=FMT_KRW)
    for j in range(n_proj):
        c = ws.cell(row=r, column=3 + j,
                    value=f'=ABS({fs_sheet}!{_pcl(j)}{fs_rows["capex"]})')
        _style(c, h_align='right', number_format=FMT_KRW)
    r += 1

    # Unlevered FCF = NOPAT + D&A - Capex
    dcf_ufcf_row = r
    _lbl(r, '비레버리지 FCF (Unlevered FCF)', bold=True)
    c = ws.cell(row=r, column=2,
                value=f'=B{dcf_nopat_row}+B{dcf_add_da_row}-B{dcf_capex_row}')
    _style(c, bold=True, fill_hex=LIGHT_GREEN, h_align='right',
           number_format=FMT_KRW)
    c.border = BOT_MED
    for j in range(n_proj):
        cl = _dcf_col(j)
        c = ws.cell(row=r, column=3 + j,
                    value=f'={cl}{dcf_nopat_row}+{cl}{dcf_add_da_row}-{cl}{dcf_capex_row}')
        _style(c, bold=True, fill_hex=LIGHT_GREEN, h_align='right',
               number_format=FMT_KRW)
        c.border = BOT_MED
    r += 2

    # Discount Factor = 1 / (1 + WACC)^year
    dcf_df_row = r
    _lbl(r, '할인계수 (Discount Factor)')
    c = ws.cell(row=r, column=2, value='')
    _style(c, fill_hex=MED_GRAY, h_align='center')
    for j in range(n_proj):
        yr_num = j + 1
        c = ws.cell(row=r, column=3 + j,
                    value=f'=1/(1+$B${wacc_ref_row})^{yr_num}')
        _style(c, h_align='right', number_format='0.0000')
    r += 1

    # PV of FCF = Unlevered FCF * Discount Factor
    dcf_pv_row = r
    _lbl(r, '현가 FCF (PV of FCF)', bold=True)
    c = ws.cell(row=r, column=2, value='')
    _style(c, fill_hex=MED_GRAY, h_align='center')
    for j in range(n_proj):
        cl = _dcf_col(j)
        c = ws.cell(row=r, column=3 + j,
                    value=f'={cl}{dcf_ufcf_row}*{cl}{dcf_df_row}')
        _style(c, bold=True, fill_hex=LIGHT_GREEN, h_align='right',
               number_format=FMT_KRW)
    r += 2

    # ── Valuation Summary ────────────────────────────────────────────────
    _write_section_header(ws, r, '기업가치 산정 (Valuation Summary)', cols=3)
    r += 1

    # Sum of PV (FCF)
    sum_pv_row = r
    _lbl(r, 'FCF 현가 합계 (Sum of PV of FCF)')
    pv_range = f'C{dcf_pv_row}:{get_column_letter(2 + n_proj)}{dcf_pv_row}'
    c = ws.cell(row=r, column=2, value=f'=SUM({pv_range})')
    _style(c, fill_hex=LIGHT_GREEN, h_align='right', number_format=FMT_KRW)
    c.border = THIN_BOX
    r += 1

    # Terminal Value = Last Year FCF * (1 + TG) / (WACC - TG)
    tv_row = r
    last_proj_col = get_column_letter(2 + n_proj)
    _lbl(r, '잔존가치 (Terminal Value)')
    c = ws.cell(row=r, column=2,
                value=f'={last_proj_col}{dcf_ufcf_row}*(1+$B${tg_row})'
                      f'/($B${wacc_ref_row}-$B${tg_row})')
    _style(c, fill_hex=LIGHT_GREEN, h_align='right', number_format=FMT_KRW)
    c.border = THIN_BOX
    r += 1

    # PV of Terminal Value
    pv_tv_row = r
    _lbl(r, '잔존가치 현가 (PV of Terminal Value)')
    c = ws.cell(row=r, column=2,
                value=f'=B{tv_row}*{last_proj_col}{dcf_df_row}')
    _style(c, fill_hex=LIGHT_GREEN, h_align='right', number_format=FMT_KRW)
    c.border = THIN_BOX
    r += 2

    # Enterprise Value
    ev_dcf_row = r
    _lbl(r, '기업가치 (Enterprise Value)', bold=True)
    c = ws.cell(row=r, column=2,
                value=f'=B{sum_pv_row}+B{pv_tv_row}')
    _style(c, fill_hex=LIGHT_BLUE, bold=True, h_align='right',
           number_format=FMT_KRW)
    c.border = BOT_MED
    c.font = _font(bold=True, size=11)
    r += 1

    # Less: Net Debt
    nd_dcf_row = r
    _lbl(r, '(-) 순부채 (Less: Net Debt)')
    c = ws.cell(row=r, column=2,
                value=f'={wacc_sheet}!B{wacc_rows["net_debt"]}')
    _style(c, fill_hex=LIGHT_BLUE, h_align='right', number_format=FMT_KRW)
    c.border = THIN_BOX
    r += 1

    # Equity Value
    eq_val_row = r
    _lbl(r, '자기자본가치 (Equity Value)', bold=True)
    c = ws.cell(row=r, column=2,
                value=f'=B{ev_dcf_row}-B{nd_dcf_row}')
    _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right',
           number_format=FMT_KRW)
    c.border = BOT_MED
    c.font = _font(bold=True, size=11)
    r += 1

    # Diluted Shares
    sh_dcf_row = r
    _lbl(r, '희석주식수 (Diluted Shares)')
    c = ws.cell(row=r, column=2,
                value=f'={wacc_sheet}!B{wacc_rows["diluted_shares"]}')
    _style(c, fill_hex=LIGHT_BLUE, h_align='right', number_format=FMT_SHARES)
    c.border = THIN_BOX
    r += 2

    # Implied Share Price
    implied_row = r
    _lbl(r, '적정주가 (Implied Share Price)', bold=True)
    c = ws.cell(row=r, column=2,
                value=f'=IF(B{sh_dcf_row}>0,B{eq_val_row}/B{sh_dcf_row},0)')
    _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right',
           number_format='#,##0')
    c.border = BOT_THICK
    c.font = _font(bold=True, size=14)
    r += 1

    # Current Price for comparison
    cur_price_dcf_row = r
    _lbl(r, '현재 주가 (Current Share Price)')
    c = ws.cell(row=r, column=2,
                value=f'={wacc_sheet}!B{wacc_rows["price"]}')
    _style(c, fill_hex=LIGHT_BLUE, h_align='right', number_format='#,##0')
    c.border = THIN_BOX
    r += 1

    # Upside / Downside %
    _lbl(r, '상승/하락 여력 (Upside/Downside %)', bold=True)
    c = ws.cell(row=r, column=2,
                value=f'=IF(B{cur_price_dcf_row}>0,'
                      f'B{implied_row}/B{cur_price_dcf_row}-1,"")')
    _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right',
           number_format='+0.0%;-0.0%')
    c.border = BOT_MED
    r += 2

    # ── Sensitivity Table ────────────────────────────────────────────────
    _write_section_header(ws, r,
        '민감도분석 (Sensitivity: WACC vs Terminal Growth)', cols=8)
    r += 1

    wacc_steps = [-0.02, -0.01, 0, 0.01, 0.02]
    tg_steps = [-0.01, -0.005, 0, 0.005, 0.01]

    # Header row
    c = ws.cell(row=r, column=1, value='WACC \\ TG')
    _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE, h_align='center')
    c.border = THIN_BOX
    for ci, tg_d in enumerate(tg_steps, 2):
        c = ws.cell(row=r, column=ci,
                    value=f'=$B${tg_row}+{tg_d}')
        _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE,
               h_align='center', number_format=FMT_PCT)
        c.border = THIN_BOX
    r += 1

    # Data rows
    for ri, w_d in enumerate(wacc_steps):
        wacc_cell_ref = f'$B${wacc_ref_row}+{w_d}'
        c = ws.cell(row=r + ri, column=1, value=f'={wacc_cell_ref}')
        _style(c, fill_hex=MED_BLUE, bold=True, font_color=WHITE,
               h_align='center', number_format=FMT_PCT)
        c.border = THIN_BOX

        for ci, tg_d in enumerate(tg_steps, 2):
            w_ref = f'($B${wacc_ref_row}+{w_d})'
            tg_ref = f'($B${tg_row}+{tg_d})'
            formula = (
                f'=IF(B{sh_dcf_row}>0,'
                f'(B{sum_pv_row}'
                f'+{last_proj_col}{dcf_ufcf_row}*(1+{tg_ref})'
                f'/({w_ref}-{tg_ref})'
                f'*{last_proj_col}{dcf_df_row}'
                f'-B{nd_dcf_row})/B{sh_dcf_row},0)'
            )
            c = ws.cell(row=r + ri, column=ci, value=formula)
            fill = YELLOW if w_d == 0 and tg_d == 0 else WHITE
            _style(c, fill_hex=fill, h_align='right', number_format='#,##0')
            c.border = THIN_BOX

    return {'wacc_ref': wacc_ref_row, 'tg': tg_row, 'implied_price': implied_row}


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 4 -- PE COMPS  (P/E 비교기업 분석)
# ─────────────────────────────────────────────────────────────────────────────

def _write_pe_comps_sheet(ws, company_info, financial_data, fs_rows):
    """Write the P/E Comparable Companies sheet."""

    def _lbl(row_num, text, bold=False, fill=None, ind=0):
        c = ws.cell(row=row_num, column=1, value='  ' * ind + text)
        _style(c, fill_hex=fill, bold=bold)
        return c

    pe_comps = financial_data.get('pe_comps', [])
    inc = financial_data['income_statement']
    years = financial_data['years']
    latest_yr = sorted(years)[-1]

    corp_name = company_info.get('corp_name', '')
    stock_code = company_info.get('stock_code', '')
    wacc_inputs = financial_data.get('wacc_inputs', {})
    current_price = (wacc_inputs.get('current_price', {}).get('price') or 0)

    _set_col_widths(ws, {1: 30, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16})

    r = 1
    title = ws.cell(row=r, column=1,
                    value=f"{corp_name} ({stock_code})"
                          f" - P/E 비교기업 분석 (PE Comparable Companies)")
    _style(title, fill_hex=DARK_BLUE, bold=True, font_color=WHITE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 2

    # Target company info
    _write_section_header(ws, r, '대상기업 (Target Company)', cols=6)
    r += 1

    _lbl(r, '현재 주가 (Current Price)')
    c = ws.cell(row=r, column=2, value=current_price)
    _style(c, fill_hex=LIGHT_BLUE, h_align='right', number_format='#,##0')
    r += 1

    target_ni = inc['net_income'].get(latest_yr) or 0
    _lbl(r, f'당기순이익 FY{latest_yr} (Net Income)')
    c = ws.cell(row=r, column=2, value=target_ni)
    _style(c, fill_hex=LIGHT_BLUE, h_align='right', number_format=FMT_KRW)
    target_ni_row = r
    r += 1

    target_shares = inc['shares_diluted'].get(latest_yr) or 0
    _lbl(r, '희석주식수 (Diluted Shares)')
    c = ws.cell(row=r, column=2, value=target_shares)
    _style(c, fill_hex=LIGHT_BLUE, h_align='right', number_format=FMT_SHARES)
    target_sh_row = r
    r += 1

    target_eps_row = r
    _lbl(r, 'EPS (주당순이익)')
    c = ws.cell(row=r, column=2,
                value=f'=IF(B{target_sh_row}>0,B{target_ni_row}/B{target_sh_row},0)')
    _style(c, fill_hex=LIGHT_GREEN, h_align='right', number_format='#,##0')
    r += 1

    target_pe_row = r
    _lbl(r, 'Trailing P/E', bold=True)
    c = ws.cell(row=r, column=2,
                value=f'=IF(B{target_eps_row}>0,{current_price}/B{target_eps_row},"")')
    _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right',
           number_format='0.0x')
    r += 2

    # ── Comparable Companies Table ───────────────────────────────────────
    _write_section_header(ws, r, '비교기업 (Comparable Companies)', cols=6)
    r += 1

    headers = ['기업명 (Company)', '시가총액 (MktCap)',
               'FY 순이익 (FY NI)', 'Trailing P/E',
               'Forward EPS', 'Forward P/E']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=ci, value=h)
        _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    r += 1

    comp_pe_start = r
    for comp in pe_comps:
        ws.cell(row=r, column=1, value=comp.get('name', '')).border = THIN_BOX

        mc = comp.get('market_cap', 0)
        c = ws.cell(row=r, column=2, value=mc if mc else '')
        _style(c, h_align='right', number_format='#,##0')
        c.border = THIN_BOX

        ni = comp.get('fy_ni')
        c = ws.cell(row=r, column=3, value=ni if ni else '')
        _style(c, h_align='right', number_format=FMT_KRW)
        c.border = THIN_BOX

        tpe = comp.get('trailing_pe')
        c = ws.cell(row=r, column=4, value=tpe if tpe else '')
        _style(c, h_align='right', number_format='0.0x')
        c.border = THIN_BOX

        feps = comp.get('forward_eps')
        c = ws.cell(row=r, column=5, value=feps if feps else '')
        _style(c, h_align='right', number_format='#,##0')
        c.border = THIN_BOX

        fpe = comp.get('forward_pe')
        c = ws.cell(row=r, column=6, value=fpe if fpe else '')
        _style(c, h_align='right', number_format='0.0x')
        c.border = THIN_BOX
        r += 1

    comp_pe_end = r - 1
    r += 1

    # ── Summary Statistics ───────────────────────────────────────────────
    _write_section_header(ws, r, '요약 통계 (Summary Statistics)', cols=6)
    r += 1

    if pe_comps:
        trailing_range = f'D{comp_pe_start}:D{comp_pe_end}'
        forward_range = f'F{comp_pe_start}:F{comp_pe_end}'

        avg_trailing_row = r
        for label, formula_t, formula_f in [
            ('평균 (Average)',  f'=AVERAGE({trailing_range})', f'=AVERAGE({forward_range})'),
            ('중앙값 (Median)', f'=MEDIAN({trailing_range})',  f'=MEDIAN({forward_range})'),
            ('최솟값 (Min)',    f'=MIN({trailing_range})',     f'=MIN({forward_range})'),
            ('최댓값 (Max)',    f'=MAX({trailing_range})',     f'=MAX({forward_range})'),
        ]:
            _lbl(r, label)
            c = ws.cell(row=r, column=4, value=formula_t)
            _style(c, fill_hex=LIGHT_GREEN, h_align='right', number_format='0.0x')
            c.border = THIN_BOX
            c = ws.cell(row=r, column=6, value=formula_f)
            _style(c, fill_hex=LIGHT_GREEN, h_align='right', number_format='0.0x')
            c.border = THIN_BOX
            r += 1

        med_trailing_row = avg_trailing_row + 1
        r += 1

        _lbl(r, '평균 P/E 기준 적정주가 (Implied Price, Avg P/E)', bold=True)
        c = ws.cell(row=r, column=2,
                    value=f'=IF(B{target_eps_row}>0,'
                          f'D{avg_trailing_row}*B{target_eps_row},"")')
        _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right',
               number_format='#,##0')
        c.border = THIN_BOX
        r += 1

        _lbl(r, '중앙값 P/E 기준 적정주가 (Implied Price, Median P/E)', bold=True)
        c = ws.cell(row=r, column=2,
                    value=f'=IF(B{target_eps_row}>0,'
                          f'D{med_trailing_row}*B{target_eps_row},"")')
        _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right',
               number_format='#,##0')
        c.border = THIN_BOX


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 5 -- DATA VALIDATION  (데이터 검증)
# ─────────────────────────────────────────────────────────────────────────────

def _run_checks(financial_data):
    """Run programmatic validation checks on the financial data."""
    years = financial_data['years']
    inc = financial_data['income_statement']
    bs = financial_data['balance_sheet']
    cf = financial_data['cash_flow']
    checks = []

    def _near(a, b, tol_pct=0.01):
        if a is None or b is None:
            return True
        if a == 0 and b == 0:
            return True
        denom = max(abs(a), abs(b), 1)
        return abs(a - b) / denom < tol_pct

    # 1. GP = Revenue - COGS
    for yr in years:
        rev = inc['revenue'].get(yr)
        cogs = inc['cogs'].get(yr)
        gp = inc['gross_profit'].get(yr)
        if rev is not None and cogs is not None and gp is not None:
            expected = rev - cogs
            ok = _near(gp, expected, 0.001)
            checks.append({
                'name': f'FY{yr} 매출총이익 = 매출액 - 매출원가 (GP = Rev - COGS)',
                'status': 'PASS' if ok else 'FAIL',
                'expected': expected, 'actual': gp,
            })

    # 2. Assets = Liabilities + Equity
    for yr in years:
        ta = bs['total_assets'].get(yr)
        tl = bs['total_liabilities'].get(yr)
        te = bs['total_equity'].get(yr)
        if ta is not None and tl is not None and te is not None:
            ok = _near(ta, tl + te, 0.001)
            checks.append({
                'name': f'FY{yr} 자산 = 부채 + 자본 (A = L + E)',
                'status': 'PASS' if ok else 'FAIL',
                'expected': ta, 'actual': tl + te,
            })

    # 3. Current Assets = sum of components (relaxed tolerance for Korean IFRS)
    for yr in years:
        tca = bs['total_current_a'].get(yr)
        if tca is not None:
            comp_sum = sum(bs[k].get(yr) or 0 for k in (
                'cash', 'st_investments', 'accounts_rec', 'inventory',
                'other_current_a'))
            ok = _near(tca, comp_sum, 0.10)  # 10% tolerance: IFRS has many sub-items
            checks.append({
                'name': f'FY{yr} 유동자산 ≈ 구성항목 합계 (CA ≈ Sum Components)',
                'status': 'PASS' if ok else 'FAIL',
                'expected': tca, 'actual': comp_sum,
            })

    # 4. FCF = Operating CF - |Capex|
    for yr in years:
        opcf = cf['operating_cf'].get(yr)
        capex = cf['capex'].get(yr)
        fcf = cf['fcf'].get(yr)
        if opcf is not None and capex is not None and fcf is not None:
            expected = opcf - abs(capex)
            ok = _near(fcf, expected, 0.001)
            checks.append({
                'name': f'FY{yr} FCF = 영업CF - |Capex|',
                'status': 'PASS' if ok else 'FAIL',
                'expected': expected, 'actual': fcf,
            })

    # 5. Net Income data availability
    ni_count = sum(1 for yr in years if inc['net_income'].get(yr) is not None)
    checks.append({
        'name': '당기순이익 데이터 존재 (Net Income data available)',
        'status': 'PASS' if ni_count >= 2 else 'FAIL',
        'detail': f'{ni_count}/{len(years)} years',
    })

    # 6. Revenue should be positive
    for yr in years:
        rev = inc['revenue'].get(yr)
        if rev is not None:
            checks.append({
                'name': f'FY{yr} 매출액 > 0 (Revenue positive)',
                'status': 'PASS' if rev > 0 else 'FAIL',
                'detail': f'{rev:,.0f}',
            })

    # 7. Total Assets should be positive
    for yr in years:
        ta = bs['total_assets'].get(yr)
        if ta is not None:
            checks.append({
                'name': f'FY{yr} 자산총계 > 0 (Total Assets positive)',
                'status': 'PASS' if ta > 0 else 'FAIL',
                'detail': f'{ta:,.0f}',
            })

    return checks


def _write_validation_sheet(ws, company_info, financial_data, fs_rows):
    """Write the Data Validation sheet.  Returns validation summary dict."""

    def _lbl(row_num, text, bold=False, fill=None, ind=0):
        c = ws.cell(row=row_num, column=1, value='  ' * ind + text)
        _style(c, fill_hex=fill, bold=bold)
        return c

    checks = _run_checks(financial_data)

    corp_name = company_info.get('corp_name', '')
    stock_code = company_info.get('stock_code', '')

    _set_col_widths(ws, {1: 55, 2: 14, 3: 20, 4: 20})

    r = 1
    title = ws.cell(row=r, column=1,
                    value=f"{corp_name} ({stock_code})"
                          f" - 데이터 검증 (Data Validation)")
    _style(title, fill_hex=DARK_BLUE, bold=True, font_color=WHITE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 2

    # Summary
    total = len(checks)
    passed = sum(1 for c in checks if c['status'] == 'PASS')
    failed = total - passed

    _lbl(r, '검증 요약 (Validation Summary)', bold=True, fill=LIGHT_BLUE)
    for ci in range(2, 5):
        ws.cell(row=r, column=ci).fill = _fill(LIGHT_BLUE)
    r += 1

    _lbl(r, '전체 검증 수 (Total Checks)')
    ws.cell(row=r, column=2, value=total)
    r += 1

    _lbl(r, '통과 (Passed)')
    c = ws.cell(row=r, column=2, value=passed)
    _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right')
    r += 1

    _lbl(r, '실패 (Failed)')
    c = ws.cell(row=r, column=2, value=failed)
    if failed > 0:
        _style(c, fill_hex=LIGHT_RED, bold=True, h_align='right')
    else:
        _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right')
    r += 1

    _lbl(r, '통과율 (Pass Rate)')
    c = ws.cell(row=r, column=2,
                value=passed / total if total > 0 else 1.0)
    _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='right',
           number_format=FMT_PCT)
    r += 2

    # Detail table
    _write_section_header(ws, r, '상세 검증 결과 (Detailed Results)', cols=4)
    r += 1

    headers = ['검증 항목 (Check)', '결과 (Status)', '기대값 (Expected)', '실제값 (Actual)']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=ci, value=h)
        _style(c, fill_hex=DARK_BLUE, bold=True, font_color=WHITE, h_align='center')
        c.border = THIN_BOX
    r += 1

    for check in checks:
        ws.cell(row=r, column=1, value=check['name']).border = THIN_BOX

        status = check['status']
        c = ws.cell(row=r, column=2, value=status)
        if status == 'PASS':
            _style(c, fill_hex=LIGHT_GREEN, bold=True, h_align='center')
        else:
            _style(c, fill_hex=LIGHT_RED, bold=True, h_align='center')
        c.border = THIN_BOX

        expected = check.get('expected')
        actual = check.get('actual')
        detail = check.get('detail', '')

        if expected is not None:
            c = ws.cell(row=r, column=3, value=expected)
            _style(c, h_align='right', number_format=FMT_KRW)
        elif detail:
            ws.cell(row=r, column=3, value=detail)
        ws.cell(row=r, column=3).border = THIN_BOX

        if actual is not None:
            c = ws.cell(row=r, column=4, value=actual)
            _style(c, h_align='right', number_format=FMT_KRW)
        ws.cell(row=r, column=4).border = THIN_BOX
        r += 1

    return {
        'total': total,
        'passed': passed,
        'failed': failed,
        'checks': checks,
    }


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def create_excel(company_info: Dict, financial_data: Dict,
                 output_file: str) -> Dict:
    """Create the full financial model workbook.

    Returns validation results dict ``{total, passed, failed, checks}``.
    """
    wb = Workbook()

    # Sheet 1: Financial Statements
    ws_fs = wb.active
    ws_fs.title = 'Financial Statements'
    fs_rows = _write_financial_statements(ws_fs, company_info, financial_data)

    # Sheet 2: WACC
    ws_wacc = wb.create_sheet('WACC')
    wacc_rows = _write_wacc_sheet(ws_wacc, company_info, financial_data, fs_rows)

    # Sheet 3: DCF Model
    ws_dcf = wb.create_sheet('DCF Model')
    _write_dcf_model(ws_dcf, company_info, financial_data, fs_rows, wacc_rows)

    # Sheet 4: PE Comps
    ws_pe = wb.create_sheet('PE Comps')
    _write_pe_comps_sheet(ws_pe, company_info, financial_data, fs_rows)

    # Sheet 5: Data Validation
    ws_val = wb.create_sheet('Data Validation')
    validation_results = _write_validation_sheet(
        ws_val, company_info, financial_data, fs_rows)

    wb.save(output_file)
    print(f"  [OK] Saved: {output_file}")

    return validation_results
